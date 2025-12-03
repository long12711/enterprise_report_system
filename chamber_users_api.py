"""
工商联用户管理 - Flask API 实现
文件: chamber_users_api.py
描述: 提供工商联用户管理的完整API实现
"""

from flask import Blueprint, request, jsonify, session, send_file
from sqlalchemy import and_, or_
from datetime import datetime
import uuid
import bcrypt
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 假设这些是已定义的模块
# from models import ChamberUser, ChamberUserLog, db
# from utils import PermissionChecker, log_operation

# 创建蓝图
chamber_users_bp = Blueprint('chamber_users', __name__, url_prefix='/api/portal/chamber')


# ============================================================
# 权限检查工具类
# ============================================================

class PermissionChecker:
    """权限检查工具"""
    
    @staticmethod
    def get_current_user():
        """获取当前登录用户"""
        user_id = session.get('user_id')
        if not user_id:
            return None
        try:
            from models import ChamberUser
            return db.session.query(ChamberUser).filter_by(id=user_id).first()
        except:
            return None
    
    @staticmethod
    def can_view_user(current_user, target_user):
        """检查是否可以查看用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以看所有用户
        if current_user.level == 'national':
            return True
        
        # 省级管理员可以看本省的用户
        if current_user.level == 'province':
            return target_user.region == current_user.region and \
                   target_user.level in ['county', 'province']
        
        # 县市级管理员只能看本县市的用户
        if current_user.level == 'county':
            return target_user.region == current_user.region and \
                   target_user.level == 'county'
        
        return False
    
    @staticmethod
    def can_edit_user(current_user, target_user):
        """检查是否可以编辑用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以编辑所有用户
        if current_user.level == 'national':
            return True
        
        # 省级管理员可以编辑本省的非管理员用户
        if current_user.level == 'province':
            if target_user.role == 'admin':
                return False
            return target_user.region == current_user.region and \
                   target_user.level in ['county', 'province']
        
        # 县市级管理员只能编辑本县市的操作员
        if current_user.level == 'county':
            if target_user.role in ['admin', 'reviewer']:
                return False
            return target_user.region == current_user.region and \
                   target_user.level == 'county'
        
        return False
    
    @staticmethod
    def can_delete_user(current_user, target_user):
        """检查是否可以删除用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以删除所有用户
        if current_user.level == 'national':
            return True
        
        # 省级管理员可以删除本省的操作员
        if current_user.level == 'province':
            if target_user.role != 'operator':
                return False
            return target_user.region == current_user.region and \
                   target_user.level in ['county', 'province']
        
        # 县市级管理员只能删除本县市的操作员
        if current_user.level == 'county':
            if target_user.role != 'operator':
                return False
            return target_user.region == current_user.region and \
                   target_user.level == 'county'
        
        return False
    
    @staticmethod
    def can_create_user(current_user, new_user_level, new_user_region):
        """检查是否可以创建用户"""
        if not current_user:
            return False
        
        # 权限层级
        level_hierarchy = {'county': 1, 'province': 2, 'national': 3}
        
        # 不能创建高于自己权限的用户
        if level_hierarchy.get(new_user_level, 0) > level_hierarchy.get(current_user.level, 0):
            return False
        
        # 检查地区权限
        if current_user.level == 'national':
            return True
        
        if current_user.level == 'province':
            return new_user_level in ['county', 'province'] and \
                   new_user_region == current_user.region
        
        if current_user.level == 'county':
            return new_user_level == 'county' and \
                   new_user_region == current_user.region
        
        return False


# ============================================================
# 辅助函数
# ============================================================

def log_operation(operator_id, target_user_id, action, old_value, new_value):
    """记录用户操作日志"""
    try:
        from models import ChamberUserLog
        log = ChamberUserLog(
            id=str(uuid.uuid4()),
            operator_id=operator_id,
            target_user_id=target_user_id,
            action=action,
            old_value=json.dumps(old_value, default=str) if old_value else None,
            new_value=json.dumps(new_value, default=str) if new_value else None
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"记录日志失败: {e}")


def hash_password(password):
    """密码加密"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password, hashed):
    """密码验证"""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def validate_password(password):
    """验证密码强度"""
    if len(password) < 8:
        return False, "密码至少8位"
    return True, "OK"


def user_to_dict(user):
    """将用户对象转换为字典"""
    return {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'real_name': user.real_name,
        'phone': user.phone,
        'level': user.level,
        'region': user.region,
        'role': user.role,
        'review_level': user.review_level,
        'department': user.department,
        'position': user.position,
        'status': user.status,
        'created_at': user.created_at.isoformat() if user.created_at else None,
        'updated_at': user.updated_at.isoformat() if user.updated_at else None,
        'remark': user.remark
    }


# ============================================================
# API 端点
# ============================================================

@chamber_users_bp.route('/users', methods=['GET'])
def get_users():
    """获取工商联用户列表"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)
        keyword = request.args.get('keyword', '')
        level = request.args.get('level', '')
        role = request.args.get('role', '')
        status = request.args.get('status', '')
        
        # 构建查询
        query = db.session.query(ChamberUser)
        
        # 权限过滤
        if current_user.level == 'county':
            query = query.filter(
                and_(
                    ChamberUser.level == 'county',
                    ChamberUser.region == current_user.region
                )
            )
        elif current_user.level == 'province':
            query = query.filter(
                and_(
                    ChamberUser.level.in_(['county', 'province']),
                    ChamberUser.region == current_user.region
                )
            )
        # national 可以看所有用户
        
        # 条件过滤
        if keyword:
            query = query.filter(
                or_(
                    ChamberUser.username.ilike(f'%{keyword}%'),
                    ChamberUser.email.ilike(f'%{keyword}%'),
                    ChamberUser.real_name.ilike(f'%{keyword}%')
                )
            )
        
        if level:
            query = query.filter(ChamberUser.level == level)
        
        if role:
            query = query.filter(ChamberUser.role == role)
        
        if status:
            query = query.filter(ChamberUser.status == status)
        
        # 排序和分页
        total = query.count()
        users = query.order_by(ChamberUser.created_at.desc()) \
                     .offset((page - 1) * limit) \
                     .limit(limit) \
                     .all()
        
        user_list = [user_to_dict(user) for user in users]
        
        return jsonify({
            'success': True,
            'data': {
                'total': total,
                'page': page,
                'limit': limit,
                'users': user_list
            }
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@chamber_users_bp.route('/users', methods=['POST'])
def create_user():
    """创建工商联用户"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        data = request.get_json()
        
        # 验证必填字段
        required_fields = ['username', 'email', 'password', 'level', 'region', 'role']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'缺少必填字段: {field}'}), 400
        
        # 权限检查
        if not PermissionChecker.can_create_user(current_user, data['level'], data['region']):
            return jsonify({'success': False, 'error': '权限不足，不能创建该用户'}), 403
        
        # 检查用户名和邮箱唯一性
        if db.session.query(ChamberUser).filter_by(username=data['username']).first():
            return jsonify({'success': False, 'error': '用户名已存在'}), 409
        
        if db.session.query(ChamberUser).filter_by(email=data['email']).first():
            return jsonify({'success': False, 'error': '邮箱已存在'}), 409
        
        # 密码验证
        password = data['password']
        valid, msg = validate_password(password)
        if not valid:
            return jsonify({'success': False, 'error': msg}), 400
        
        # 创建用户
        user = ChamberUser(
            id=str(uuid.uuid4()),
            username=data['username'],
            email=data['email'],
            password=hash_password(password),
            real_name=data.get('real_name'),
            phone=data.get('phone'),
            level=data['level'],
            region=data['region'],
            role=data['role'],
            review_level=data.get('review_level'),
            department=data.get('department'),
            position=data.get('position'),
            status=data.get('status', 'pending'),
            remark=data.get('remark'),
            created_by=current_user.id
        )
        
        db.session.add(user)
        db.session.commit()
        
        # 记录操作日志
        log_operation(current_user.id, user.id, 'create', None, user_to_dict(user))
        
        return jsonify({
            'success': True,
            'data': {
                'id': user.id,
                'username': user.username,
                'message': '用户创建成功'
            }
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['GET'])
def get_user(user_id):
    """获取用户详情"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        target_user = db.session.query(ChamberUser).filter_by(id=user_id).first()
        if not target_user:
            return jsonify({'success': False, 'error': '用户不存在'}), 404
        
        # 权限检查
        if not PermissionChecker.can_view_user(current_user, target_user):
            return jsonify({'success': False, 'error': '权限不足'}), 403
        
        return jsonify({
            'success': True,
            'data': user_to_dict(target_user)
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['PUT'])
def update_user(user_id):
    """更新工商联用户"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        target_user = db.session.query(ChamberUser).filter_by(id=user_id).first()
        if not target_user:
            return jsonify({'success': False, 'error': '用户不存在'}), 404
        
        # 权限检查
        if not PermissionChecker.can_edit_user(current_user, target_user):
            return jsonify({'success': False, 'error': '权限不足'}), 403
        
        data = request.get_json()
        old_value = user_to_dict(target_user)
        
        # 更新字段
        if 'email' in data:
            # 检查邮箱唯一性
            existing = db.session.query(ChamberUser).filter(
                and_(
                    ChamberUser.email == data['email'],
                    ChamberUser.id != user_id
                )
            ).first()
            if existing:
                return jsonify({'success': False, 'error': '邮箱已被使用'}), 409
            target_user.email = data['email']
        
        if 'password' in data and data['password']:
            valid, msg = validate_password(data['password'])
            if not valid:
                return jsonify({'success': False, 'error': msg}), 400
            target_user.password = hash_password(data['password'])
        
        if 'real_name' in data:
            target_user.real_name = data['real_name']
        
        if 'phone' in data:
            target_user.phone = data['phone']
        
        if 'role' in data:
            target_user.role = data['role']
        
        if 'review_level' in data:
            target_user.review_level = data['review_level']
        
        if 'department' in data:
            target_user.department = data['department']
        
        if 'position' in data:
            target_user.position = data['position']
        
        if 'status' in data:
            target_user.status = data['status']
        
        if 'remark' in data:
            target_user.remark = data['remark']
        
        target_user.updated_at = datetime.now()
        db.session.commit()
        
        # 记录操作日志
        log_operation(current_user.id, user_id, 'update', old_value, user_to_dict(target_user))
        
        return jsonify({
            'success': True,
            'data': {
                'id': user_id,
                'message': '用户更新成功'
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    """删除工商联用户"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        target_user = db.session.query(ChamberUser).filter_by(id=user_id).first()
        if not target_user:
            return jsonify({'success': False, 'error': '用户不存在'}), 404
        
        # 权限检查
        if not PermissionChecker.can_delete_user(current_user, target_user):
            return jsonify({'success': False, 'error': '权限不足'}), 403
        
        old_value = user_to_dict(target_user)
        db.session.delete(target_user)
        db.session.commit()
        
        # 记录操作日志
        log_operation(current_user.id, user_id, 'delete', old_value, None)
        
        return jsonify({'success': True, 'message': '用户删除成功'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@chamber_users_bp.route('/users/export', methods=['GET'])
def export_users():
    """导出用户列表为Excel"""
    try:
        from models import ChamberUser
        
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'success': False, 'error': '未登录'}), 401
        
        # 获取查询参数
        keyword = request.args.get('keyword', '')
        level = request.args.get('level', '')
        role = request.args.get('role', '')
        status = request.args.get('status', '')
        
        # 构建查询（同列表接口）
        query = db.session.query(ChamberUser)
        
        if current_user.level == 'county':
            query = query.filter(
                and_(
                    ChamberUser.level == 'county',
                    ChamberUser.region == current_user.region
                )
            )
        elif current_user.level == 'province':
            query = query.filter(
                and_(
                    ChamberUser.level.in_(['county', 'province']),
                    ChamberUser.region == current_user.region
                )
            )
        
        if keyword:
            query = query.filter(
                or_(
                    ChamberUser.username.ilike(f'%{keyword}%'),
                    ChamberUser.email.ilike(f'%{keyword}%'),
                    ChamberUser.real_name.ilike(f'%{keyword}%')
                )
            )
        
        if level:
            query = query.filter(ChamberUser.level == level)
        
        if role:
            query = query.filter(ChamberUser.role == role)
        
        if status:
            query = query.filter(ChamberUser.status == status)
        
        users = query.all()
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '工商联用户'
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # 写入表头
        headers = ['用户名', '邮箱', '真实姓名', '手机', '层级', '地区', '角色', 
                   '审核权限', '部门', '职位', '状态', '创建时间', '备注']
        
        header_fill = PatternFill(start_color='4338CA', end_color='4338CA', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        level_map = {'county': '县市级', 'province': '省级', 'national': '全联'}
        role_map = {'admin': '管理员', 'reviewer': '审核员', 'operator': '操作员'}
        status_map = {'active': '激活', 'inactive': '停用', 'pending': '待审核'}
        
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1).value = user.username
            ws.cell(row=row, column=2).value = user.email
            ws.cell(row=row, column=3).value = user.real_name or ''
            ws.cell(row=row, column=4).value = user.phone or ''
            ws.cell(row=row, column=5).value = level_map.get(user.level, '')
            ws.cell(row=row, column=6).value = user.region or ''
            ws.cell(row=row, column=7).value = role_map.get(user.role, '')
            ws.cell(row=row, column=8).value = user.review_level or ''
            ws.cell(row=row, column=9).value = user.department or ''
            ws.cell(row=row, column=10).value = user.position or ''
            ws.cell(row=row, column=11).value = status_map.get(user.status, '')
            ws.cell(row=row, column=12).value = user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
            ws.cell(row=row, column=13).value = user.remark or ''
        
        # 返回文件
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'工商联用户_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# 错误处理
# ============================================================

@chamber_users_bp.errorhandler(404)
def not_found(error):
    return jsonify({'success': False, 'error': '资源不存在'}), 404


@chamber_users_bp.errorhandler(500)
def internal_error(error):
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


# ============================================================
# 蓝图注册
# ============================================================

def register_chamber_users_api(app):
    """注册工商联用户管理API"""
    app.register_blueprint(chamber_users_bp)


# 使用示例:
# from flask import Flask
# app = Flask(__name__)
# register_chamber_users_api(app)

