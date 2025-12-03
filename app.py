"""
企业现代制度评价系统 - 主应用
功能：
1. 在线问卷填写
2. Excel文件上传
3. 自评报告生成
4. 邮件/短信发送
5. 管理员功能
"""

from flask import Flask, request, render_template, jsonify, send_file, url_for, redirect, session
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime, timedelta
import uuid
import json
from pathlib import Path
import threading
import hashlib
from functools import wraps
# from report_generator import ReportGenerator  # 已废弃，使用 enterprise_report_generator
from report_generator.notification_service import NotificationService
from survey_engine.questionnaire_submission_manager import QuestionnaireSubmissionManager
from report_engine import EnterpriseReportGenerator, PDFReportGenerator, ProfessionalReportGenerator, ComprehensiveAnalysisGenerator
from report_generator.nankai_report_generator import NankaiReportGenerator
from survey_generator.nankai_scoring_engine import NankaiScoringEngine
# from score_calculator import ScoreCalculator  # 避免依赖空指标文件导致初始化失败
from user_types_config_final import (
    get_all_user_types, 
    get_user_levels, 
    get_questionnaire_config
)
# from nankai_indicator_loader import load_questions_by_level, map_user_to_level  # 已迁移到 survey_engine
from survey_engine.services.loader import SurveyLoader

# 问卷指标文件配置
try:
    from survey_config import INDICATOR_FILE
except ImportError:
    # 如果配置文件不存在，使用默认值
    import os
    INDICATOR_FILE_CANDIDATES = ["nankai_indicators.xlsx", "测试问卷.xlsx", "指标体系.xlsx"]
    INDICATOR_FILE = next((f for f in INDICATOR_FILE_CANDIDATES if os.path.exists(f)), None)
    if INDICATOR_FILE:
        print(f"[配置] 使用指标文件: {INDICATOR_FILE}")
    else:
        print("[警告] 未找到指标文件，南开问卷功能将不可用")

app = Flask(__name__)
# 注册专家门户蓝图
from expert_portal import ui_bp as expert_ui_bp, api_bp as expert_api_bp
app.register_blueprint(expert_ui_bp, url_prefix='/portal/expert')
app.register_blueprint(expert_api_bp, url_prefix='/api/portal/expert')
# 注册问卷引擎蓝图（提供 /api/get_questions、/api/health 等）
from survey_engine import api_bp as survey_api_bp
app.register_blueprint(survey_api_bp, url_prefix='/api/survey')
# 注册指标体系管理蓝图
from indicator_management import indicator_management_bp
app.register_blueprint(indicator_management_bp)

# 指标管理页面路由
@app.route('/admin/indicator-management')
def admin_indicator_management():
    """指标体系管理页面"""
    return render_template('admin_indicator_management.html')

app.config['UPLOAD_FOLDER'] = 'storage/uploads'
app.config['REPORT_FOLDER'] = 'storage/reports'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SUBMISSIONS_FOLDER'] = 'storage/submissions'
app.config['SPECIAL_FOLDER'] = 'storage/special_submissions'
app.config['SECRET_KEY'] = 'your-secret-key-here-change-this-in-production'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# 指标文件管理：统一由 SurveyLoader 管理
_survey_loader = SurveyLoader()

# 确保文件夹存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['REPORT_FOLDER'], exist_ok=True)
os.makedirs(app.config['SUBMISSIONS_FOLDER'], exist_ok=True)

# 初始化服务
# report_generator = ReportGenerator()  # 已废弃，使用 enterprise_report_generator
notification_service = NotificationService()
# 临时禁用submission_manager以避免ScoreCalculator初始化失败
try:
    submission_manager = QuestionnaireSubmissionManager(storage_folder=app.config['SUBMISSIONS_FOLDER'])
except Exception as e:
    print(f"[WARN] QuestionnaireSubmissionManager初始化失败: {e}")
    submission_manager = None
enterprise_report_generator = EnterpriseReportGenerator()
pdf_report_generator = PDFReportGenerator()

# 新增：专业报告生成器（集中由 report_engine 提供）
professional_report_generator = ProfessionalReportGenerator() if 'ProfessionalReportGenerator' in globals() and ProfessionalReportGenerator else None

# 新增：综合分析报告生成器（集中由 report_engine 提供）
comprehensive_analysis_generator = ComprehensiveAnalysisGenerator() if 'ComprehensiveAnalysisGenerator' in globals() and ComprehensiveAnalysisGenerator else None

# 新增：南开问卷自评报告生成器
nankai_report_generator = NankaiReportGenerator(indicator_file=INDICATOR_FILE) if INDICATOR_FILE else None

# 新增：南开问卷评分引擎（使用标准评分细则）
nankai_scoring_engine = NankaiScoringEngine('survey_generator/南开大学-现代企业制度指数评价体系初稿2025.10.22（单独指标）(1).xlsx')

# 存储处理状态
processing_status = {}

# 管理员账户 (生产环境应该使用数据库)
ADMIN_USERS = {
    'admin': hashlib.sha256('admin123'.encode()).hexdigest()  # 默认密码: admin123
}

# 认证装饰器
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return jsonify({'error': '未授权访问'}), 401
        return f(*args, **kwargs)
    return decorated_function

# 读取账户配置（简易内置账号）
try:
    with open('config.json', 'r', encoding='utf-8') as _cf:
        _conf = json.load(_cf)
        ACCOUNTS = _conf.get('accounts', {})
except Exception:
    ACCOUNTS = {}

DEFAULT_ACCOUNTS = {
    'enterprise': [{'username': 'ent', 'password': 'ent123', 'display_name': '企业示例', 'default_level': 'advanced'}],
    'chamber_of_commerce': [{'username': 'chamber', 'password': 'chamber123', 'display_name': '工商联示例', 'default_level': 'advanced'}],
    'expert': [{'username': 'expert', 'password': 'expert123', 'display_name': '专家示例', 'default_level': 'senior'}],
}

ROLE_PORTAL = {
    'enterprise': '/portal/enterprise',
    'chamber_of_commerce': '/portal/chamber',
    'expert': '/portal/expert',
}

LEVEL_RANK = {'beginner': 1, 'intermediate': 2, 'advanced': 3}
LEVEL_LABEL_ENTERPRISE = {'beginner': '初级', 'intermediate': '中级', 'advanced': '高级'}
LEVEL_LABEL_CHAMBER = {'beginner': '市级', 'intermediate': '省级', 'advanced': '国家级'}

# 等级记录文件（简单KV）。示例：{"enterprise:ent":"advanced"}
USER_LEVELS_FILE = 'user_levels.json'

def _load_user_levels():
    try:
        with open(USER_LEVELS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _save_user_levels(d):
    try:
        with open(USER_LEVELS_FILE, 'w', encoding='utf-8') as f:
            json.dump(d, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _get_default_level(role, username):
    # 从config账户的default_level或内置默认
    accs = ACCOUNTS.get(role) or DEFAULT_ACCOUNTS.get(role, [])
    for a in accs:
        if a.get('username') == username:
            lvl = a.get('default_level')
            if lvl:
                return lvl
    if role == 'expert':
        return 'senior'
    return 'advanced'

def _get_record_level(role, username):
    d = _load_user_levels()
    return d.get(f"{role}:{username}")

def _set_record_level(role, username, level):
    d = _load_user_levels()
    d[f"{role}:{username}"] = level
    _save_user_levels(d)

def _next_enterprise_level(cur):
    mapping = {'beginner': 'intermediate', 'intermediate': 'advanced'}
    return mapping.get(cur)


def role_required(role):
    def wrapper(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if session.get('role') != role:
                return redirect('/')
            return f(*args, **kwargs)
        return inner
    return wrapper

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json(force=True)
        role = data.get('role')
        username = data.get('username', '')
        password = data.get('password', '')
        remember = data.get('remember', False)

        if role not in ['enterprise', 'chamber_of_commerce', 'expert']:
            return jsonify({'success': False, 'error': '角色无效'}), 400

        candidates = ACCOUNTS.get(role) or DEFAULT_ACCOUNTS.get(role, [])
        ok = next((u for u in candidates if u.get('username') == username and u.get('password') == password), None)
        if not ok:
            return jsonify({'success': False, 'error': '账户或密码错误'}), 401

        # 读取或生成用户等级
        record_level = _get_record_level(role, username)
        if not record_level:
            record_level = _get_default_level(role, username)
            _set_record_level(role, username, record_level)

        session['logged_in'] = True
        session['role'] = role
        session['username'] = username
        session['display_name'] = ok.get('display_name', username)
        session['user_level'] = record_level
        session.permanent = bool(remember)

        return jsonify({'success': True, 'redirect': ROLE_PORTAL.get(role)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def api_logout():
    for k in ['logged_in','role','username','display_name','user_level']:
        session.pop(k, None)
    return jsonify({'success': True})

@app.route('/api/session')
def api_session():
    return jsonify({
        'logged_in': bool(session.get('logged_in')),
        'role': session.get('role'),
        'username': session.get('username'),
        'display_name': session.get('display_name'),
        'user_level': session.get('user_level'),
    })

# 门户页面
@app.route('/portal/enterprise')
@role_required('enterprise')
def portal_enterprise():
    return render_template('portal_enterprise.html')

@app.route('/portal/chamber')
@role_required('chamber_of_commerce')
def portal_chamber():
    return render_template('portal_chamber.html')


# ======= 工商联审核/专家匹配 API =======

def _list_submission_jsons(limit=200):
    base = app.config['SUBMISSIONS_FOLDER']
    items = []
    if not os.path.exists(base):
        return items
    for fn in os.listdir(base):
        if fn.endswith('.json'):
            p = os.path.join(base, fn)
            try:
                items.append((os.path.getmtime(p), p))
            except:  # noqa
                pass
    items.sort(reverse=True)
    return [p for _, p in items[:limit]]


def _compute_score_from_excel(xlsx_path):
    try:
        df = pd.read_excel(xlsx_path, sheet_name='问卷')
        # 兼容列名
        if '答案' not in df.columns and '答案/清单选择' in df.columns:
            df = df.rename(columns={'答案/清单选择': '答案'})
        if '指标类型' not in df.columns and '问题类型' in df.columns:
            df = df.rename(columns={'问题类型': '指标类型'})
        # 基础分
        def to_float(v, d=0.0):
            try:
                if pd.isna(v):
                    return d
                return float(v)
            except:
                return d
        # 快速路径：存在计算分数列则直接求和
        if '计算分数' in df.columns:
            total_score = df['计算分数'].apply(lambda x: to_float(x, 0.0)).sum()
        else:
            # 简化版打分逻辑
            eff_map = {'很有效':1.0,'比较有效':0.8,'一般':0.6,'不太有效':0.3,'完全无效':0.0}
            total_score = 0.0
            for _, r in df.iterrows():
                base = to_float(r.get('分值'), 0.0)
                ans = r.get('答案')
                qtype = str(r.get('指标类型') or '')
                # 尝试JSON
                if isinstance(ans, str) and ans.strip().startswith('{'):
                    try:
                        jr = json.loads(ans)
                        total_score += to_float(jr.get('score'), 0.0)
                        continue
                    except:  # noqa
                        pass
                ans = str(ans or '').strip()
                if '否决' in qtype or '调节' in qtype:
                    total_score += base if ans == '否' else 0.0
                elif '合规' in qtype:
                    total_score += base if ans == '是' else 0.0
                elif '有效' in qtype:
                    total_score += base * eff_map.get(ans, 0.0)
                else:
                    total_score += base if ans in ['是','有','已建立','已设立','已制定'] else 0.0
        # 满分：仅统计正分题
        max_possible = df['分值'].apply(lambda x: to_float(x, 0.0)).apply(lambda x: x if x>0 else 0).sum()
        pct = (total_score/max_possible*100.0) if max_possible>0 else 0.0
        return {'total': float(total_score), 'max': float(max_possible), 'percentage': float(pct)}
    except Exception as e:
        return {'total': 0.0, 'max': 0.0, 'percentage': 0.0, 'error': str(e)}


def _load_submission(json_path):
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # 推导excel文件
        base = os.path.basename(json_path)
        xlsx = os.path.join('submissions', base.replace('submission_', '问卷_').replace('.json', '.xlsx'))
        return data, xlsx
    except Exception as e:
        return None, None


@app.route('/api/portal/chamber/all-reports')
@role_required('chamber_of_commerce')
def api_chamber_all_reports():
    """获取所有企业报告"""
    try:
        reports_folder = app.config['REPORT_FOLDER']
        reports = []

        if os.path.exists(reports_folder):
            for filename in os.listdir(reports_folder):
                if filename.endswith(('.docx', '.pdf')):
                    filepath = os.path.join(reports_folder, filename)
                    stat = os.stat(filepath)

                    # 从文件名提取企业名称
                    parts = filename.split('_')
                    enterprise_name = parts[1] if '专业报告' in parts[0] and len(parts) > 2 else parts[0]

                    reports.append({
                        'filename': filename,
                        'enterprise_name': enterprise_name,
                        'type': 'Word' if filename.endswith('.docx') else 'PDF',
                        'file_size': stat.st_size,
                        'created_time': stat.st_mtime
                    })

        reports.sort(key=lambda x: x['created_time'], reverse=True)

        return jsonify({'success': True, 'reports': reports})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/portal/chamber/approve-upgrade', methods=['POST'])
@role_required('chamber_of_commerce')
def api_chamber_approve_upgrade():
    """批准企业升级"""
    try:
        data = request.get_json()
        username = data.get('username')
        new_level = data.get('new_level')

        if not username or not new_level:
            return jsonify({'success': False, 'error': '参数不完整'}), 400

        _set_record_level('enterprise', username, new_level)
        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/portal/chamber/reviews')
@role_required('chamber_of_commerce')
def api_chamber_reviews():
    try:
        reviewer_level = session.get('user_level', 'advanced')
        level = request.args.get('level', 'auto')
        target = reviewer_level if level=='auto' else level
        res_items = []
        for jp in _list_submission_jsons():
            data, xlsx = _load_submission(jp)
            if not data or not os.path.exists(xlsx):
                continue
            cur_level = data.get('user_level','advanced')
            # 仅显示与审核级别相邻或以下的：示例策略
            if LEVEL_RANK.get(cur_level,3) > LEVEL_RANK.get(target,3):
                continue
            sc = _compute_score_from_excel(xlsx)
            pct = sc.get('percentage',0.0)
            # 升级阈值：初->中 70%，中->高 80%
            upgrade_to = None
            eligible = False
            if cur_level=='beginner' and pct>=70:
                upgrade_to = 'intermediate'; eligible=True
            elif cur_level=='intermediate' and pct>=80:
                upgrade_to = 'advanced'; eligible=True
            # 关联原始问卷Excel与最新报告
            enterprise_name = data.get('enterprise_info',{}).get('企业名称','')
            excel_file = os.path.basename(xlsx) if os.path.exists(xlsx) else None
            report_file = None
            reports_folder = app.config['REPORT_FOLDER']
            if os.path.exists(reports_folder):
                found_reports = []
                for f in os.listdir(reports_folder):
                    if enterprise_name and enterprise_name in f and f.endswith(('.docx', '.pdf')):
                        fp = os.path.join(reports_folder, f)
                        found_reports.append((os.path.getmtime(fp), f))
                if found_reports:
                    found_reports.sort(key=lambda x: x[0], reverse=True)
                    report_file = found_reports[0][1]

            res_items.append({
                'enterprise_name': enterprise_name,
                'current_level': cur_level,
                'score_percentage': pct,
                'eligible': eligible,
                'upgrade_to': upgrade_to,
                'username': data.get('username'),
                'submit_time': os.path.getmtime(jp),
                'submit_time_text': datetime.fromtimestamp(os.path.getmtime(jp)).strftime('%Y-%m-%d %H:%M'),
                'excel_file': excel_file,
                'report_file': report_file,
                'next_level': _next_enterprise_level(cur_level)
            })
        return jsonify({'success': True, 'items': res_items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """
    处理Excel文件上传
    """
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '只支持Excel文件(.xlsx, .xls)'}), 400

    try:
        # 保存上传的文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"upload_{timestamp}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # 读取Excel文件
        df = pd.read_excel(filepath)

        # 验证必需的列
        required_columns = ['企业名称', '联系人', '联系人邮箱', '联系人手机']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return jsonify({
                'error': f'Excel文件缺少必需的列: {", ".join(missing_columns)}',
                'columns': df.columns.tolist()
            }), 400

        # 返回预览数据
        preview_data = df.head(5).to_dict('records')
        total_records = len(df)

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'total_records': total_records,
            'preview': preview_data,
            'columns': df.columns.tolist()
        })

    except Exception as e:
        return jsonify({'error': f'文件处理失败: {str(e)}'}), 500


@app.route('/process', methods=['POST'])
def process_file():
    """
    处理Excel文件并生成报告
    """
    data = request.get_json()
    filepath = data.get('filepath')
    send_email = data.get('send_email', True)
    send_sms = data.get('send_sms', False)

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 400

    # 生成任务ID
    task_id = datetime.now().strftime('%Y%m%d%H%M%S')
    processing_status[task_id] = {
        'status': 'processing',
        'total': 0,
        'processed': 0,
        'success': 0,
        'failed': 0,
        'errors': []
    }

    # 异步处理
    thread = threading.Thread(
        target=process_excel_async,
        args=(task_id, filepath, send_email, send_sms)
    )
    thread.start()

    return jsonify({
        'success': True,
        'task_id': task_id,
        'message': '开始处理，请稍后查看进度'
    })


def process_excel_async(task_id, filepath, send_email, send_sms):
    """
    异步处理Excel文件
    """
    with app.app_context():
        try:
            # 读取Excel
            df = pd.read_excel(filepath)
            total = len(df)
            processing_status[task_id]['total'] = total
            processing_status[task_id]['status'] = 'processing'

            # 逐行处理
            for idx, row in df.iterrows():
                try:
                    # 提取企业信息
                    enterprise_data = row.to_dict()

                    # 生成报告
                    report_path = enterprise_report_generator.generate_report(
                        enterprise_data,
                        output_folder=app.config['REPORT_FOLDER']
                    )

                    # 生成报告访问URL
                    report_filename = os.path.basename(report_path)
                    # 直接构造URL，避免在异步线程中使用url_for
                    report_url = f"http://localhost:5000/download/{report_filename}"

                    # 发送通知
                    contact_name = enterprise_data.get('联系人', '')
                    enterprise_name = enterprise_data.get('企业名称', '')
                    email = enterprise_data.get('联系人邮箱', '')
                    phone = enterprise_data.get('联系人手机', '')

                    # 发送邮件
                    if send_email and email:
                        email_sent = notification_service.send_email(
                            to_email=email,
                            enterprise_name=enterprise_name,
                            contact_name=contact_name,
                            report_url=report_url,
                            attachment_path=report_path
                        )
                        if not email_sent:
                            processing_status[task_id]['errors'].append(
                                f'{enterprise_name}: 邮件发送失败'
                            )

                    # 发送短信
                    if send_sms and phone:
                        sms_sent = notification_service.send_sms(
                            phone=phone,
                            enterprise_name=enterprise_name,
                            report_url=report_url
                        )
                        if not sms_sent:
                            processing_status[task_id]['errors'].append(
                                f'{enterprise_name}: 短信发送失败'
                            )

                    processing_status[task_id]['processed'] += 1
                    processing_status[task_id]['success'] += 1

                except Exception as e:
                    processing_status[task_id]['processed'] += 1
                    processing_status[task_id]['failed'] += 1
                    processing_status[task_id]['errors'].append(
                        f'第{idx+1}行处理失败: {str(e)}'
                    )

            processing_status[task_id]['status'] = 'completed'

        except Exception as e:
            processing_status[task_id]['status'] = 'failed'
            processing_status[task_id]['errors'].append(f'整体处理失败: {str(e)}')


@app.route('/status/<task_id>')
def get_status(task_id):
    """
    获取处理状态
    """
    if task_id not in processing_status:
        return jsonify({'error': '任务不存在'}), 404

    return jsonify(processing_status[task_id])


@app.route('/download/submission/<filename>')
def download_submission(filename):
    """下载历史问卷Excel文件"""
    filepath = os.path.join('submissions', filename)

    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404

    return send_file(filepath, as_attachment=True)


@app.route('/download/<filename>')
def download_report(filename):
    """
    下载报告
    """
    filepath = os.path.join(app.config['REPORT_FOLDER'], filename)

    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404

    return send_file(filepath, as_attachment=True)


@app.route('/test_email', methods=['POST'])
def test_email():
    """
    测试邮件发送
    """
    data = request.get_json()
    email = data.get('email')

    if not email:
        return jsonify({'error': '请提供邮箱地址'}), 400

    try:
        success = notification_service.send_test_email(email)
        if success:
            return jsonify({'success': True, 'message': '测试邮件发送成功'})
        else:
            return jsonify({'success': False, 'message': '测试邮件发送失败'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/test_sms', methods=['POST'])
def test_sms():
    """
    测试短信发送
    """
    data = request.get_json()
    phone = data.get('phone')

    if not phone:
        return jsonify({'error': '请提供手机号码'}), 400

    try:
        success = notification_service.send_test_sms(phone)
        if success:
            return jsonify({'success': True, 'message': '测试短信发送成功'})
        else:
            return jsonify({'success': False, 'message': '测试短信发送失败'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============ 在线问卷填写功能 ============

@app.route('/api/user-types')
def get_user_types_api():
    """
    获取所有用户类型和分级
    """
    try:
        user_types = get_all_user_types()
        
        # 添加每个用户类型的分级信息
        result = []
        for user_type in user_types:
            levels = get_user_levels(user_type['value'])
            result.append({
                'value': user_type['value'],
                'name': user_type['name'],
                'description': user_type['description'],
                'levels': levels
            })
        
        return jsonify({
            'success': True,
            'user_types': result
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/questionnaire')
def questionnaire():
    """在线问卷填写页面"""
    return render_template('questionnaire.html')


@app.route('/success')
def success_page():
    """问卷提交成功页面"""
    return render_template('success.html')


@app.route('/api/get_questions')
def get_questions():
    """
    获取问卷题目（兼容旧路径）。内部委托给 SurveyLoader，保持与 /api/survey/get_questions 一致。
    """
    try:
        # 获取查询参数
        user_type = request.args.get('user_type')
        user_level = request.args.get('user_level')
        excel_level = request.args.get('excel_level')  # 可选：beginner|intermediate|advanced

        level_key = _survey_loader.resolve_level_key(user_type, user_level, excel_level)

        questions = _survey_loader.get_questions(level_key)
        source_file = _survey_loader.nk_excel_path

        return jsonify({
            'success': True,
            'questions': questions,
            'user_type': user_type,
            'user_level': user_level,
            'excel_level': level_key,
            'source_file': source_file,
            'total_questions': len(questions)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'获取问卷题目失败: {str(e)}'
        }), 500


@app.route('/api/health')
def health():
    """运行状态与指标文件健康检查（兼容旧路径，委托给 SurveyLoader）"""
    try:
        path = _survey_loader.nk_excel_path
        exists = bool(path) and os.path.exists(path)
        result = {
            'success': True,
            'nk_excel_path': path,
            'file_exists': exists,
        }
        if exists:
            try:
                overview = {}
                for lv in ['beginner','intermediate','advanced']:
                    qs = _survey_loader.get_questions(lv)
                    overview[lv] = len(qs)
                result['overview'] = overview
            except Exception as e:
                result['overview_error'] = str(e)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/indicators/overview')
def indicators_overview():
    """返回各分级题目数量与示例题目（前3条）（兼容旧路径，委托 SurveyLoader）"""
    try:
        path = _survey_loader.nk_excel_path
        if not path or not os.path.exists(path):
            return jsonify({'success': False, 'error': '指标文件不存在'}), 500
        data = {}
        for lv in ['beginner','intermediate','advanced']:
            qs = _survey_loader.get_questions(lv)
            data[lv] = {
                'count': len(qs),
                'samples': qs[:3]
            }
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/nankai-survey')
def nankai_survey_entrance():
    """南开问卷系统入口页面"""
    return render_template('nankai_survey_entrance.html')


@app.route('/nankai-questionnaire')
def nankai_questionnaire_page():
    """南开问卷预览页面（旧版）"""
    level = request.args.get('level', '初级')
    return render_template('nankai_questionnaire.html', selected_level=level)


@app.route('/nankai-questionnaire-fill')
def nankai_questionnaire_fill_page():
    """南开问卷填写页面"""
    try:
        import pandas as pd
        
        level = request.args.get('level', '初级')
        excel_path = INDICATOR_FILE
        
        if not excel_path or not os.path.exists(excel_path):
            return f"<h1>指标文件不存在</h1><p>请确保以下文件之一存在：</p><ul><li>nankai_indicators.xlsx</li><li>测试问卷.xlsx</li></ul>", 404
        
        # 读取Excel数据
        df = pd.read_excel(excel_path, sheet_name=level)
        
        # 构建问卷数据
        questions = []
        
        current_category = ''
        for idx, row in df.iterrows():
            q_id = int(row['序号']) if pd.notna(row['序号']) else idx + 1
            
            # 读取评分准则（题目内容）
            question_text = ''
            for col_name in ['评分准则', '评分标准', '打分标准', '题目']:
                if col_name in df.columns and pd.notna(row[col_name]):
                    question_text = str(row[col_name])
                    break
            
            # 如果没有评分准则，使用三级指标作为题目
            if not question_text:
                question_text = str(row['三级指标']) if pd.notna(row['三级指标']) else ''
            
            # 解析选项列
            options_col = None
            for col_name in ['选项', '打分标准']:
                if col_name in df.columns:
                    options_col = col_name
                    break
            
            options_text = str(row[options_col]) if options_col and pd.notna(row[options_col]) else ''
            options_list = []
            if options_text:
                for line in options_text.split('\n'):
                    line = line.strip()
                    if line and (line.startswith('A.') or line.startswith('B.') or line.startswith('C.')):
                        options_list.append(line)
            
            if not options_list:
                options_list = ['A. 已完成', 'B. 部分完成', 'C. 未完成']
            
            # 读取分值
            score = 0
            score_value = '1'
            if '分值' in df.columns and pd.notna(row['分值']):
                try:
                    score_str = str(row['分值'])
                    score_value = score_str  # 保留原始分值字符串（如"0-2"）
                    # 处理范围分值，如"0-2"，取最大值
                    if '-' in score_str:
                        score = float(score_str.split('-')[1])
                    else:
                        score = float(score_str)
                except:
                    score = 1  # 默认1分
                    score_value = '1'
            
            # 读取评分准则（详细的评分规则）
            scoring_rule = ''
            for col_name in ['评分准则', '评分规则', '评分说明']:
                if col_name in df.columns and pd.notna(row[col_name]):
                    scoring_rule = str(row[col_name])
                    break
            
            # 读取佐证材料要求
            evidence = ''
            if '佐证材料' in df.columns and pd.notna(row['佐证材料']):
                evidence = str(row['佐证材料'])
            
            questions.append({
                'id': q_id,
                'category': str(row['一级指标']) if pd.notna(row['一级指标']) else current_category,
                'level1': str(row['一级指标']) if pd.notna(row['一级指标']) else '',
                'level2': str(row['二级指标']) if pd.notna(row['二级指标']) else '',
                'level3': str(row['三级指标']) if pd.notna(row['三级指标']) else '',
                'indicator': question_text,
                'options': options_list,
                'score': score,
                'score_value': score_value,  # 新增：分值字符串
                'scoring_rule': scoring_rule,  # 新增：评分准则
                'evidence': evidence  # 新增：佐证材料
            })
            
            if pd.notna(row['一级指标']):
                current_category = str(row['一级指标'])
        
        return render_template(
            'nankai_questionnaire_fill.html',
            level_name=level,
            questions=questions
        )
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 加载问卷失败: {error_detail}")
        return f"<h1>加载问卷失败</h1><pre>{error_detail}</pre>", 500


@app.route('/api/nankai/questionnaire/<level>')
def get_nankai_questionnaire(level):
    """获取南开问卷数据（初级/中级/高级）"""
    try:
        from survey_generator.nankai_excel_generator import NankaiExcelGenerator
        
        excel_path = INDICATOR_FILE
        
        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': '指标文件不存在，请联系管理员'}), 404
        
        generator = NankaiExcelGenerator(excel_path)
        
        # 映射级别名称
        level_map = {
            'beginner': '初级',
            'intermediate': '中级',
            'advanced': '高级',
            '初级': '初级',
            '中级': '中级',
            '高级': '高级'
        }
        
        chinese_level = level_map.get(level)
        if not chinese_level:
            return jsonify({'success': False, 'error': '无效的级别'}), 400
        
        # 获取数据
        df = generator.data[chinese_level].copy()
        
        # 添加选项列
        df['选项'] = df['打分标准'].apply(generator._parse_scoring_standard)
        
        # 转换为字典列表
        questions = []
        for _, row in df.iterrows():
            questions.append({
                '序号': int(row['序号']) if pd.notna(row['序号']) else 0,
                '一级指标': str(row['一级指标']) if pd.notna(row['一级指标']) else '',
                '二级指标': str(row['二级指标']) if pd.notna(row['二级指标']) else '',
                '三级指标': str(row['三级指标']) if pd.notna(row['三级指标']) else '',
                '打分标准': str(row['打分标准']) if pd.notna(row['打分标准']) else '',
                '选项': str(row['选项']) if pd.notna(row['选项']) else '',
                '佐证材料': str(row['佐证材料']) if pd.notna(row['佐证材料']) else ''
            })
        
        return jsonify({
            'success': True,
            'level': chinese_level,
            'total': len(questions),
            'questions': questions
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nankai/submit', methods=['POST'])
def submit_nankai_questionnaire():
    """提交南开问卷并生成评分和报告"""
    try:
        # 处理multipart/form-data格式的数据
        enterprise_info = {}
        answers = {}
        partial_details = {}  # 存储部分完成的详细说明
        files = {}
        level = request.form.get('level', '初级')
        
        # 解析表单数据
        for key in request.form:
            value = request.form[key]
            
            # 企业基本信息
            if key in ['enterprise_name', 'contact_person', 'contact_phone', 'contact_email',
                      'main_business', 'enterprise_scale', 'establishment_years',
                      'annual_revenue', 'rd_investment', 'rd_ratio']:
                enterprise_info[key] = value
            
            # 问卷答案
            elif key.startswith('q_'):
                question_id = key[2:]  # 去掉 'q_' 前缀
                answers[question_id] = value
            
            # 部分完成的详细说明
            elif key.startswith('partial_detail_'):
                question_id = key[15:]  # 去掉 'partial_detail_' 前缀
                if value.strip():  # 只保存非空的说明
                    partial_details[question_id] = value.strip()
        
        # 处理上传的文件
        for key in request.files:
            if key.startswith('evidence_'):
                question_id = key[9:]  # 去掉 'evidence_' 前缀
                uploaded_files = request.files.getlist(key)
                if uploaded_files:
                    files[question_id] = [f.filename for f in uploaded_files if f.filename]
        
        if not enterprise_info or not answers:
            return jsonify({'success': False, 'error': '数据格式错误'}), 400
        
        # 使用南开评分引擎计算得分（基于评分细则）
        try:
            score_result = nankai_scoring_engine.calculate_score(
                level=level,
                answers=answers,
                partial_details=partial_details
            )
            
            total_score = score_result['total_score']
            max_score = score_result['max_score']
            final_score = score_result['percentage']
            score_details = score_result['details']
            
        except Exception as e:
            print(f"[ERROR] 使用评分引擎计算失败，回退到简化评分: {e}")
            import traceback
            traceback.print_exc()
            
            # 回退到简化评分逻辑
            excel_path = INDICATOR_FILE
            if not excel_path or not os.path.exists(excel_path):
                return jsonify({'success': False, 'error': '指标文件不存在'}), 500
            
            df = pd.read_excel(excel_path, sheet_name=level)
            
            # 构建题目分值映射
            question_scores = {}
            for idx, row in df.iterrows():
                q_id = str(int(row['序号'])) if pd.notna(row['序号']) else str(idx + 1)
                score = 0
                if '分值' in df.columns and pd.notna(row['分值']):
                    try:
                        score_str = str(row['分值'])
                        if '-' in score_str:
                            score = float(score_str.split('-')[1])
                        else:
                            score = float(score_str)
                    except:
                        score = 1
                question_scores[q_id] = score
            
            # 计算得分
            total_score = 0
            max_score = sum(question_scores.values())
            score_details = {}
            
            for question_id, answer in answers.items():
                q_score = question_scores.get(question_id, 1)
                option_letter = answer[0] if answer else 'C'
                
                if option_letter == 'A':
                    earned_score = q_score
                elif option_letter == 'B':
                    earned_score = q_score * 0.8
                else:
                    earned_score = 0
                
                total_score += earned_score
                score_details[question_id] = {
                    'max_score': q_score,
                    'earned_score': earned_score,
                    'answer': answer
                }
            
            final_score = (total_score / max_score * 100) if max_score > 0 else 0
        
        # 保存提交数据
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        enterprise_name = enterprise_info.get('enterprise_name', '未知企业')
        
        submission_data = {
            'timestamp': timestamp,
            'level': level,
            'enterprise_info': enterprise_info,
            'answers': answers,
            'partial_details': partial_details,  # 保存部分完成的详细说明
            'files': files,
            'score': {
                'total_score': round(total_score, 2),
                'max_score': round(max_score, 2),
                'percentage': round(final_score, 2),
                'details': score_details
            },
            'submitted_at': datetime.now().isoformat()
        }
        
        # 保存到文件
        os.makedirs('storage/nankai_submissions', exist_ok=True)
        submission_file = f'storage/nankai_submissions/submission_{enterprise_name}_{timestamp}.json'
        with open(submission_file, 'w', encoding='utf-8') as f:
            json.dump(submission_data, f, ensure_ascii=False, indent=2)
        
        # 生成提交ID用于结果页面
        submission_id = f"{enterprise_name}_{timestamp}"
        
        # 异步生成报告并发送邮件
        thread = threading.Thread(
            target=generate_and_send_nankai_report_async,
            args=(submission_file, enterprise_info, level)
        )
        thread.start()
        
        return jsonify({
            'success': True,
            'score': round(final_score, 2),
            'total_score': round(total_score, 2),
            'max_score': round(max_score, 2),
            'message': '问卷提交成功，自评报告将通过邮件发送',
            'submission_id': submission_id,
            'submission_file': submission_file
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 提交问卷失败: {error_detail}")
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_and_send_nankai_report_async(submission_file, enterprise_info, level):
    """
    异步生成南开问卷自评报告并发送邮件
    """
    with app.app_context():
        try:
            enterprise_name = enterprise_info.get('enterprise_name', '企业')
            email = enterprise_info.get('contact_email', '')
            contact_name = enterprise_info.get('contact_person', '')
            
            print(f"\n[INFO] 开始为 {enterprise_name} 生成南开自评报告...")
            
            # 生成自评报告
            if nankai_report_generator:
                report_path = nankai_report_generator.generate_report(submission_file)
                print(f"[OK] 南开自评报告已生成: {report_path}")
                
                # 发送邮件
                if email:
                    print(f"[INFO] 正在发送邮件到 {email}...")
                    
                    email_sent = notification_service.send_email(
                        to_email=email,
                        enterprise_name=enterprise_name,
                        contact_name=contact_name,
                        report_url='',  # 通过附件发送，不需要URL
                        attachment_path=report_path
                    )
                    
                    if email_sent:
                        print(f"[SUCCESS] 南开自评报告邮件发送成功到 {email}")
                    else:
                        print(f"[ERROR] 南开自评报告邮件发送失败到 {email}")
                else:
                    print(f"[WARN] {enterprise_name} 未提供邮箱地址，跳过邮件发送")
            else:
                print(f"[ERROR] 南开报告生成器未初始化")
                
        except Exception as e:
            print(f"[ERROR] 生成和发送南开自评报告失败: {e}")
            import traceback
            traceback.print_exc()


@app.route('/api/nankai/download/<submission_id>')
def download_nankai_submission(submission_id):
    """下载南开问卷填报数据为Excel格式"""
    try:
        # 查找对应的JSON文件
        json_file = f'storage/nankai_submissions/submission_{submission_id}.json'
        
        if not os.path.exists(json_file):
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        
        # 读取JSON数据
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 从Excel读取题目信息
        level = data.get('level', '初级')
        excel_path = INDICATOR_FILE
        
        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': '指标文件不存在'}), 500
        
        df = pd.read_excel(excel_path, sheet_name=level)
        
        # 构建输出数据
        output_data = []
        answers = data.get('answers', {})
        partial_details = data.get('partial_details', {})
        score_details = data.get('score', {}).get('details', {})
        
        for idx, row in df.iterrows():
            q_id = str(int(row['序号'])) if pd.notna(row['序号']) else str(idx + 1)
            
            # 获取答案
            answer = answers.get(q_id, '')
            
            # 获取部分完成的详细说明
            partial_detail = partial_details.get(q_id, '')
            
            # 获取得分信息
            score_info = score_details.get(q_id, {})
            max_score = score_info.get('max_score', 0)
            earned_score = score_info.get('earned_score', 0)
            
            output_row = {
                '序号': q_id,
                '一级指标': str(row['一级指标']) if pd.notna(row['一级指标']) else '',
                '二级指标': str(row['二级指标']) if pd.notna(row['二级指标']) else '',
                '三级指标': str(row['三级指标']) if pd.notna(row['三级指标']) else '',
                '题目': str(row['题目']) if pd.notna(row['题目']) else '',
                '分值': max_score,
                '填报答案': answer,
                '部分完成说明': partial_detail,
                '得分': earned_score,
                '评分标准': str(row['评分准则']) if pd.notna(row['评分准则']) else (
                    str(row['评分标准']) if pd.notna(row.get('评分标准', '')) else (
                        str(row['打分标准']) if pd.notna(row.get('打分标准', '')) else ''
                    )
                )
            }
            output_data.append(output_row)
        
        # 创建DataFrame
        output_df = pd.DataFrame(output_data)
        
        # 添加企业信息和总分信息
        enterprise_info = data.get('enterprise_info', {})
        score_info = data.get('score', {})
        
        # 生成Excel文件
        output_dir = 'storage/nankai_downloads'
        os.makedirs(output_dir, exist_ok=True)
        
        enterprise_name = enterprise_info.get('enterprise_name', '企业')
        timestamp = data.get('timestamp', datetime.now().strftime('%Y%m%d_%H%M%S'))
        output_filename = f'问卷填报数据_{enterprise_name}_{timestamp}.xlsx'
        output_path = os.path.join(output_dir, output_filename)
        
        # 使用openpyxl创建格式化的Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = f'{level}问卷填报数据'
        
        # 添加标题行
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f'南开大学现代企业制度指数评价问卷 - {level}'
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.row_dimensions[1].height = 30
        
        # 添加企业信息
        info_row = 2
        ws.merge_cells(f'A{info_row}:J{info_row}')
        info_cell = ws[f'A{info_row}']
        info_text = f"企业名称：{enterprise_info.get('enterprise_name', '')}  |  " \
                   f"联系人：{enterprise_info.get('contact_person', '')}  |  " \
                   f"联系电话：{enterprise_info.get('contact_phone', '')}  |  " \
                   f"提交时间：{data.get('submitted_at', '')}"
        info_cell.value = info_text
        info_cell.font = Font(size=10)
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        info_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws.row_dimensions[info_row].height = 25
        
        # 添加得分信息
        score_row = 3
        ws.merge_cells(f'A{score_row}:J{score_row}')
        score_cell = ws[f'A{score_row}']
        score_text = f"百分制得分：{score_info.get('percentage', 0):.2f}分  |  " \
                    f"实际得分：{score_info.get('total_score', 0):.2f}分  |  " \
                    f"满分：{score_info.get('max_score', 0):.2f}分"
        score_cell.value = score_text
        score_cell.font = Font(size=11, bold=True, color='2E5090')
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        score_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.row_dimensions[score_row].height = 25
        
        # 添加表头
        header_row = 5
        headers = ['序号', '一级指标', '二级指标', '三级指标', '题目', '分值', '填报答案', '部分完成说明', '得分', '评分标准']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        
        ws.row_dimensions[header_row].height = 30
        
        # 添加数据
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for r_idx, row_data in enumerate(dataframe_to_rows(output_df, index=False, header=False), header_row + 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # 根据列设置对齐方式
                if c_idx in [1, 6, 9]:  # 序号、分值、得分 - 居中
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in [5, 7, 8, 10]:  # 题目、答案、说明、评分标准 - 左对齐
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 设置列宽
        column_widths = {
            'A': 8,   # 序号
            'B': 15,  # 一级指标
            'C': 15,  # 二级指标
            'D': 20,  # 三级指标
            'E': 40,  # 题目
            'F': 8,   # 分值
            'G': 20,  # 填报答案
            'H': 30,  # 部分完成说明
            'I': 8,   # 得分
            'J': 30   # 评分标准
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 保存文件
        wb.save(output_path)
        
        # 发送文件
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 下载问卷数据失败: {error_detail}")
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/nankai-result')
def nankai_result_page():
    """南开问卷结果页面"""
    score = request.args.get('score', '0')
    total_score = request.args.get('total_score', '0')
    max_score = request.args.get('max_score', '100')
    level = request.args.get('level', '初级')
    enterprise = request.args.get('enterprise', '企业')
    return render_template('nankai_result.html',
                         score=score,
                         total_score=total_score,
                         max_score=max_score,
                         level=level,
                         enterprise=enterprise)


@app.route('/api/nankai/generate-report/<submission_id>')
def generate_nankai_report(submission_id):
    """生成南开问卷自评报告"""
    try:
        if not nankai_report_generator:
            return jsonify({'success': False, 'error': '报告生成器未初始化'}), 500
        
        # 查找对应的JSON文件
        json_file = f'storage/nankai_submissions/submission_{submission_id}.json'
        
        if not os.path.exists(json_file):
            return jsonify({'success': False, 'error': '提交记录不存在'}), 404
        
        # 生成报告
        report_path = nankai_report_generator.generate_report(json_file)
        
        # 返回下载链接
        report_filename = os.path.basename(report_path)
        
        return jsonify({
            'success': True,
            'report_path': report_path,
            'report_filename': report_filename,
            'download_url': f'/api/nankai/download-report/{report_filename}'
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 生成自评报告失败: {error_detail}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nankai/download-report/<filename>')
def download_nankai_report(filename):
    """下载南开问卷自评报告（保留旧接口兼容性）"""
    try:
        from urllib.parse import unquote
        # 解码URL中的中文字符
        filename = unquote(filename, encoding='utf-8')
        
        report_path = os.path.join('storage/nankai_reports', filename)
        
        if not os.path.exists(report_path):
            return jsonify({'success': False, 'error': '报告文件不存在'}), 404
        
        return send_file(
            report_path,
            as_attachment=True,
            attachment_filename=filename,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 下载报告失败: {error_detail}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/nankai/download-report-by-id/<submission_id>')
def download_nankai_report_by_id(submission_id):
    """
    通过submission_id下载报告（推荐使用，避免文件名编码问题）
    """
    try:
        reports_dir = 'storage/nankai_reports'
        
        if not os.path.exists(reports_dir):
            return jsonify({'success': False, 'error': '报告目录不存在'}), 404
        
        # 在报告目录中查找包含submission_id的文件
        report_file = None
        for filename in os.listdir(reports_dir):
            if submission_id in filename and filename.endswith('.docx'):
                report_file = filename
                break
        
        if not report_file:
            return jsonify({'success': False, 'error': '报告文件不存在'}), 404
        
        report_path = os.path.join(reports_dir, report_file)
        
        # 使用安全的文件名
        safe_filename = f'自评报告_{submission_id}.docx'
        
        return send_file(
            report_path,
            as_attachment=True,
            attachment_filename=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"[ERROR] 下载报告失败: {error_detail}")
        return jsonify({'success': False, 'error': str(e)}), 500


def filter_questions_by_user_type(questions, user_type, user_level):
    """
    根据用户类型和分级过滤问卷题目
    """
    try:
        # 获取问卷配置
        config = get_questionnaire_config(user_type, user_level)
        
        if not config:
            # 如果没有配置，返回所有题目
            return questions
        
        filtered_questions = []
        
        for question in questions:
            # 检查问题类型是否在允许的类型中
            if question['question_type'] not in config['question_types']:
                continue
            
            # 检查适用对象是否匹配
            applicable = question['applicable_enterprises']
            if applicable not in config['applicable_enterprises'] and '所有企业' not in config['applicable_enterprises']:
                continue
            
            # 如果配置中指定了重点领域，则只包含这些领域的题目
            if not config['include_all'] and 'focus_areas' in config:
                if question['level1'] not in config['focus_areas']:
                    continue
            
            filtered_questions.append(question)
        
        return filtered_questions
    
    except Exception as e:
        print(f"过滤问卷题目出错: {e}")
        return questions


@app.route('/api/submit_questionnaire', methods=['POST'])
def submit_questionnaire():
    """
    接收问卷提交
    """
    try:
        data = request.get_json()
        print("--- Received submission data ---")
        print(json.dumps(data, indent=2, ensure_ascii=False))
        print("-----------------------------")
        print("--- Received submission data ---")
        print(json.dumps(data, indent=2, ensure_ascii=False))
        print("-----------------------------")

        if not data or 'enterprise_info' not in data or 'answers' not in data:
            return jsonify({'error': '数据格式错误'}), 400

        # 从会话中获取用户类型和分级，更可靠
        user_type = session.get('role')
        user_level = session.get('user_level')

        if not user_type or not user_level:
            return jsonify({'error': '无法从会话中获取用户身份，请重新登录'}), 401

        # 将会话中的用户身份信息补充到提交数据中，确保后续流程能正确使用
        data['user_type'] = user_type
        data['user_level'] = user_level
        data['username'] = session.get('username')  # 关联用户名

        # 保存提交数据
        result = submission_manager.save_submission(data)

        enterprise_name = result['enterprise_name']
        excel_path = result['excel_path']

        # 异步生成报告并发送邮件
        thread = threading.Thread(
            target=generate_and_send_report_async,
            args=(excel_path, data['enterprise_info'], user_type, user_level)
        )
        thread.start()

        return jsonify({
            'success': True,
            'message': '问卷提交成功，报告将通过邮件发送',
            'enterprise_name': enterprise_name,
            'user_type': user_type,
            'user_level': user_level
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'提交失败: {str(e)}'
        }), 500


def generate_and_send_report_async(questionnaire_path, enterprise_info, user_type=None, user_level=None):
    """
    异步生成报告并发送邮件（现在生成专业版报告）
    """
    with app.app_context():
        try:
            enterprise_name = enterprise_info.get('企业名称', '')
            email = enterprise_info.get('联系人邮箱', '')
            contact_name = enterprise_info.get('联系人姓名', '')
            
            # 记录用户类型和分级信息
            if user_type and user_level:
                print(f"[INFO] 用户类型: {user_type}, 分级: {user_level}")

            print(f"\n[INFO] 开始为 {enterprise_name} 生成报告...")

            # 生成专业版Word报告（叙述性风格，适合对外发送）
            if professional_report_generator:
                word_report_path = professional_report_generator.generate_report(questionnaire_path)
                print(f"[OK] 专业版报告已生成: {word_report_path}")
            else:
                # 如果专业版不可用，使用原版
                word_report_path = enterprise_report_generator.generate_report(questionnaire_path)
                print(f"[OK] Word报告已生成: {word_report_path}")

            # 生成PDF报告
            pdf_report_path = pdf_report_generator.generate_report(questionnaire_path)
            print(f"[OK] PDF报告已生成: {pdf_report_path}")

            # 发送邮件
            if email:
                print(f"[INFO] 正在发送邮件到 {email}...")

                email_sent = notification_service.send_email(
                    to_email=email,
                    enterprise_name=enterprise_name,
                    contact_name=contact_name,
                    report_url='',
                    attachment_path=word_report_path  # 通过邮件附件发送Word报告
                )

                if email_sent:
                    print(f"[SUCCESS] 邮件发送成功到 {email}")
                else:
                    print(f"[ERROR] 邮件发送失败到 {email}")
            else:
                print(f"[WARN] {enterprise_name} 未提供邮箱地址，跳过邮件发送")

        except Exception as e:
            print(f"[ERROR] 生成和发送报告失败: {e}")
            import traceback
            traceback.print_exc()


# ============ 管理员功能 ============

@app.route('/admin/login')
def admin_login_page():
    """管理员登录页面"""
    return render_template('admin_login.html')


@app.route('/admin/dashboard')
def admin_dashboard():
    """管理员仪表板"""
    if 'admin_logged_in' not in session:
        return redirect('/admin/login')
    return render_template('admin_dashboard.html')


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """管理员登录API"""
    try:
        data = request.get_json()
        username = data.get('username', '')
        password = data.get('password', '')

        # 验证用户名和密码
        password_hash = hashlib.sha256(password.encode()).hexdigest()

        if username in ADMIN_USERS and ADMIN_USERS[username] == password_hash:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            session.permanent = True

            return jsonify({
                'success': True,
                'message': '登录成功',
                'token': 'dummy-token'  # 实际应该生成JWT token
            })
        else:
            return jsonify({
                'success': False,
                'error': '用户名或密码错误'
            }), 401

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    """管理员登出"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    return jsonify({'success': True})


@app.route('/api/admin/dashboard-stats')
@admin_required
def admin_dashboard_stats():
    """获取仪表板统计数据"""
    try:
        submissions = submission_manager.get_all_submissions()

        # 统计今日提交
        today = datetime.now().date()
        today_submissions = sum(
            1 for sub in submissions
            if datetime.fromtimestamp(sub['submit_time']).date() == today
        )

        # 统计报告数量
        reports_folder = app.config['REPORT_FOLDER']
        total_reports = len([f for f in os.listdir(reports_folder) if f.endswith(('.docx', '.pdf'))]) if os.path.exists(reports_folder) else 0

        return jsonify({
            'success': True,
            'stats': {
                'total_submissions': len(submissions),
                'total_reports': total_reports,
                'today_submissions': today_submissions,
                'emails_sent': len(submissions)  # 简化统计
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/recent-submissions')
@admin_required
def admin_recent_submissions():
    """获取最近提交"""
    try:
        limit = int(request.args.get('limit', 10))
        submissions = submission_manager.get_all_submissions()

        # 按时间排序，取最新的
        sorted_submissions = sorted(
            submissions,
            key=lambda x: x['submit_time'],
            reverse=True
        )[:limit]

        return jsonify({
            'success': True,
            'submissions': sorted_submissions
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/submissions')
@admin_required
def admin_all_submissions():
    """获取所有提交"""
    try:
        submissions = submission_manager.get_all_submissions()

        # 添加额外信息
        for sub in submissions:
            # 读取提交详情
            try:
                submission_data = submission_manager.get_submission_by_filename(sub['filename'])
                enterprise_info = submission_data.get('enterprise_info', {})
                sub['contact_name'] = enterprise_info.get('联系人姓名', '')
                sub['email'] = enterprise_info.get('联系人邮箱', '')
            except:
                sub['contact_name'] = ''
                sub['email'] = ''

        return jsonify({
            'success': True,
            'submissions': submissions
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/submission/<filename>')
@admin_required
def admin_submission_detail(filename):
    """获取提交详情"""
    try:
        submission_data = submission_manager.get_submission_by_filename(filename)

        return jsonify({
            'success': True,
            'submission': submission_data
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/generate-report', methods=['POST'])
@admin_required
def admin_generate_report():
    """管理员生成报告"""
    try:
        data = request.get_json()
        filename = data.get('filename')

        if not filename:
            return jsonify({'success': False, 'error': '缺少文件名'}), 400

        # 获取对应的Excel文件
        excel_path = os.path.join('submissions', filename.replace('.json', '.xlsx').replace('submission_', '问卷_'))

        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': 'Excel文件不存在'}), 404

        # 生成报告
        word_report = enterprise_report_generator.generate_report(excel_path)
        pdf_report = pdf_report_generator.generate_report(excel_path)

        return jsonify({
            'success': True,
            'word_report': word_report,
            'pdf_report': pdf_report
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/send-email', methods=['POST'])
@admin_required
def admin_send_email():
    """管理员发送邮件"""
    try:
        data = request.get_json()
        filename = data.get('filename')

        if not filename:
            return jsonify({'success': False, 'error': '缺少文件名'}), 400

        # 获取提交数据
        submission_data = submission_manager.get_submission_by_filename(filename)
        enterprise_info = submission_data['enterprise_info']

        # 查找报告文件
        enterprise_name = enterprise_info.get('企业名称', '')
        reports_folder = app.config['REPORT_FOLDER']

        # 查找最新的报告
        word_report = None
        for f in os.listdir(reports_folder):
            if enterprise_name in f and f.endswith('.docx'):
                word_report = os.path.join(reports_folder, f)
                break

        if not word_report:
            return jsonify({'success': False, 'error': '未找到报告文件'}), 404

        # 发送邮件
        email = enterprise_info.get('联系人邮箱', '')
        contact_name = enterprise_info.get('联系人姓名', '')

        success = notification_service.send_email(
            to_email=email,
            enterprise_name=enterprise_name,
            contact_name=contact_name,
            report_url='',
            attachment_path=word_report
        )

        return jsonify({
            'success': success,
            'message': '邮件发送成功' if success else '邮件发送失败'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/reports')
@admin_required
def admin_reports():
    """获取所有报告"""
    try:
        reports_folder = app.config['REPORT_FOLDER']
        reports = []

        if os.path.exists(reports_folder):
            for filename in os.listdir(reports_folder):
                if filename.endswith(('.docx', '.pdf')):
                    filepath = os.path.join(reports_folder, filename)
                    stat = os.stat(filepath)

                    # 从文件名提取企业名称
                    enterprise_name = filename.split('_')[0] if '_' in filename else filename

                    reports.append({
                        'filename': filename,
                        'enterprise_name': enterprise_name,
                        'type': 'Word' if filename.endswith('.docx') else 'PDF',
                        'file_size': stat.st_size,
                        'created_time': stat.st_mtime
                    })

        # 按创建时间倒序排序
        reports.sort(key=lambda x: x['created_time'], reverse=True)

        return jsonify({
            'success': True,
            'reports': reports
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/report/<filename>', methods=['DELETE'])
@admin_required
def admin_delete_report(filename):
    """删除报告"""
    try:
        filepath = os.path.join(app.config['REPORT_FOLDER'], filename)

        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': '文件不存在'}), 404

        os.remove(filepath)

        return jsonify({
            'success': True,
            'message': '删除成功'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/batch-generate', methods=['POST'])
@admin_required
def admin_batch_generate():
    """批量生成报告"""
    try:
        data = request.get_json()
        filenames = data.get('filenames', [])
        generate_word = data.get('generate_word', True)
        generate_pdf = data.get('generate_pdf', True)
        send_email = data.get('send_email', False)

        if not filenames:
            return jsonify({'success': False, 'error': '未选择任何文件'}), 400

        # 创建任务ID
        task_id = datetime.now().strftime('%Y%m%d%H%M%S')
        processing_status[task_id] = {
            'status': 'processing',
            'total': len(filenames),
            'processed': 0,
            'success': 0,
            'failed': 0,
            'errors': []
        }

        # 异步处理
        thread = threading.Thread(
            target=batch_generate_async,
            args=(task_id, filenames, generate_word, generate_pdf, send_email)
        )
        thread.start()

        return jsonify({
            'success': True,
            'task_id': task_id
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def batch_generate_async(task_id, filenames, generate_word, generate_pdf, send_email_flag):
    """异步批量生成报告"""
    with app.app_context():
        for filename in filenames:
            try:
                # 获取Excel文件路径
                excel_path = os.path.join('submissions', filename.replace('.json', '.xlsx').replace('submission_', '问卷_'))

                if not os.path.exists(excel_path):
                    processing_status[task_id]['processed'] += 1
                    processing_status[task_id]['failed'] += 1
                    processing_status[task_id]['errors'].append(f'{filename}: Excel文件不存在')
                    continue

                # 获取企业信息
                submission_data = submission_manager.get_submission_by_filename(filename)
                enterprise_info = submission_data['enterprise_info']

                # 生成报告
                if generate_word:
                    enterprise_report_generator.generate_report(excel_path)

                if generate_pdf:
                    pdf_report_generator.generate_report(excel_path)

                # 发送邮件
                if send_email_flag:
                    email = enterprise_info.get('联系人邮箱', '')
                    if email:
                        notification_service.send_email(
                            to_email=email,
                            enterprise_name=enterprise_info.get('企业名称', ''),
                            contact_name=enterprise_info.get('联系人姓名', ''),
                            report_url='',
                            attachment_path=excel_path.replace('.xlsx', '_自评报告.docx')
                        )

                processing_status[task_id]['processed'] += 1
                processing_status[task_id]['success'] += 1

            except Exception as e:
                processing_status[task_id]['processed'] += 1
                processing_status[task_id]['failed'] += 1
                processing_status[task_id]['errors'].append(f'{filename}: {str(e)}')

        processing_status[task_id]['status'] = 'completed'


@app.route('/api/admin/batch-status/<task_id>')
@admin_required
def admin_batch_status(task_id):
    """获取批量任务状态"""
    if task_id not in processing_status:
        return jsonify({'success': False, 'error': '任务不存在'}), 404

    return jsonify({
        'success': True,
        'data': processing_status[task_id]
    })


@app.route('/api/admin/generate-professional-report', methods=['POST'])
@admin_required
def admin_generate_professional_report():
    """生成专业版报告（叙述性风格）"""
    if not professional_report_generator:
        return jsonify({'success': False, 'error': '专业报告生成器未加载'}), 500

    try:
        data = request.get_json()
        filename = data.get('filename')

        if not filename:
            return jsonify({'success': False, 'error': '缺少文件名'}), 400

        # 获取Excel文件路径
        excel_path = os.path.join('submissions', filename.replace('.json', '.xlsx').replace('submission_', '问卷_'))

        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': 'Excel文件不存在'}), 404

        # 生成专业版报告
        report_path = professional_report_generator.generate_report(excel_path)

        return jsonify({
            'success': True,
            'report_path': report_path
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/generate-comprehensive-report', methods=['POST'])
@admin_required
def admin_generate_comprehensive_report():
    """生成综合分析报告"""
    if not comprehensive_analysis_generator:
        return jsonify({'success': False, 'error': '综合分析生成器未加载'}), 500

    try:
        data = request.get_json()
        filenames = data.get('filenames', [])

        if not filenames:
            return jsonify({'success': False, 'error': '未选择任何文件'}), 400

        # 获取所有Excel文件路径
        excel_files = []
        for filename in filenames:
            excel_path = os.path.join('submissions', filename.replace('.json', '.xlsx').replace('submission_', '问卷_'))
            if os.path.exists(excel_path):
                excel_files.append(excel_path)

        if not excel_files:
            return jsonify({'success': False, 'error': '未找到有效的Excel文件'}), 404

        # 生成综合报告
        report_path = comprehensive_analysis_generator.generate_comprehensive_report(excel_files)

        return jsonify({
            'success': True,
            'report_path': report_path,
            'enterprise_count': len(excel_files)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ======= 突出专项申请 后端API =======
SPECIAL_FOLDER = 'special_submissions'
os.makedirs(SPECIAL_FOLDER, exist_ok=True)


def _special_json_path(sid):
    return os.path.join(SPECIAL_FOLDER, f'{sid}.json')


def _special_dir(sid):
    d = os.path.join(SPECIAL_FOLDER, sid)
    os.makedirs(d, exist_ok=True)
    return d


def _read_special(sid):
    try:
        with open(_special_json_path(sid), 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def _write_special(data):
    with open(_special_json_path(data['id']), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.route('/api/special/apply', methods=['POST'])
@role_required('enterprise')
def api_special_apply():
    try:
        title = request.form.get('title', '').strip()
        desc = request.form.get('desc', '').strip()
        if not title:
            return jsonify({'success': False, 'error': '缺少标题'}), 400
        sid = uuid.uuid4().hex[:12]
        owner = session.get('username') or 'enterprise'
        display_name = session.get('display_name') or owner
        level = session.get('user_level', 'advanced')
        ts = datetime.now().timestamp()
        record = {
            'id': sid,
            'owner': owner,
            'enterprise': display_name,
            'level': level,
            'title': title,
            'desc': desc,
            'files': [],
            'status': 'pending',
            'remark': '',
            'time': ts
        }
        # 保存附件
        upload_dir = _special_dir(sid)
        files = request.files.getlist('files')
        for f in files:
            if not f.filename:
                continue
            sf = secure_filename(f.filename) or (uuid.uuid4().hex + '.bin')
            save_path = os.path.join(upload_dir, sf)
            f.save(save_path)
            record['files'].append(sf)
        _write_special(record)
        return jsonify({'success': True, 'id': sid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/special/my')
@role_required('enterprise')
def api_special_my():
    try:
        owner = session.get('username')
        items = []
        for fn in os.listdir(SPECIAL_FOLDER):
            if fn.endswith('.json'):
                sid = fn[:-5]
                rec = _read_special(sid)
                if rec and rec.get('owner') == owner:
                    rec['time_text'] = datetime.fromtimestamp(rec['time']).strftime('%Y-%m-%d %H:%M')
                    items.append(rec)
        items.sort(key=lambda x: x['time'], reverse=True)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/special/list')
@role_required('chamber_of_commerce')
def api_special_list():
    try:
        reviewer_level = session.get('user_level', 'advanced')
        level = request.args.get('level', 'auto')
        target = reviewer_level if level == 'auto' else level
        items = []
        for fn in os.listdir(SPECIAL_FOLDER):
            if fn.endswith('.json'):
                sid = fn[:-5]
                rec = _read_special(sid)
                if rec and rec.get('level') == target:
                    rec['time_text'] = datetime.fromtimestamp(rec['time']).strftime('%Y-%m-%d %H:%M')
                    items.append(rec)
        items.sort(key=lambda x: x['time'], reverse=True)
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/special/download/<sid>/<path:filename>')
@role_required('chamber_of_commerce')
def api_special_download(sid, filename):
    try:
        rec = _read_special(sid)
        if not rec:
            return jsonify({'success': False, 'error': '申请不存在'}), 404
        # 权限：同级别工商联方可下载
        if rec.get('level') != session.get('user_level'):
            return jsonify({'success': False, 'error': '无权下载该附件（级别不匹配）'}), 403
        fpath = os.path.join(_special_dir(sid), filename)
        if not os.path.exists(fpath):
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        return send_file(fpath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/special/review', methods=['POST'])
@role_required('chamber_of_commerce')
def api_special_review():
    try:
        data = request.get_json(force=True)
        sid = data.get('id')
        action = data.get('action')  # approve/reject
        remark = data.get('remark', '')
        rec = _read_special(sid)
        if not rec:
            return jsonify({'success': False, 'error': '申请不存在'}), 404
        # 仅审核与自身级别一致的申请
        if rec.get('level') != session.get('user_level'):
            return jsonify({'success': False, 'error': '无权审核（级别不匹配）'}), 403
        if action == 'approve':
            rec['status'] = 'approved'
            # 联动升级：企业当前等级 -> 下一等级（beginner->intermediate，intermediate->advanced）
            owner_username = rec.get('owner')
            cur_level = _get_record_level('enterprise', owner_username) or rec.get('level')
            nxt = _next_enterprise_level(cur_level)
            if nxt:
                _set_record_level('enterprise', owner_username, nxt)
                rec['upgraded_to'] = nxt
        elif action == 'reject':
            rec['status'] = 'rejected'
        else:
            return jsonify({'success': False, 'error': '无效操作'}), 400
        rec['remark'] = remark
        rec['reviewer'] = session.get('username')
        rec['review_time'] = datetime.now().timestamp()
        _write_special(rec)
        return jsonify({'success': True, 'upgraded_to': rec.get('upgraded_to')})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




@app.route('/api/enterprise/history')
@role_required('enterprise')
def api_enterprise_history():
    """获取当前企业的问卷提交历史，并关联得分和报告"""
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'error': '无法获取用户信息，请重新登录'}), 401

        submissions_folder = 'submissions'
        reports_folder = app.config['REPORT_FOLDER']
        history_items = []

        all_submissions = submission_manager.get_all_submissions()
        enterprise_submissions = [s for s in all_submissions if s.get('username') == username]

        for sub in enterprise_submissions:
            try:
                submission_data = submission_manager.get_submission_by_filename(sub['filename'])
                excel_path = sub['filepath'].replace('submission_', '问卷_').replace('.json', '.xlsx')
                
                if not os.path.exists(excel_path):
                    continue

                # 计算得分
                score_info = _compute_score_from_excel(excel_path)

                # 查找关联的报告
                report_filename = None
                # 从提交数据中获取企业名称，用于匹配报告文件
                enterprise_name = submission_data.get('enterprise_info', {}).get('企业名称', '')

                if os.path.exists(reports_folder):
                    # 报告文件名可能包含时间戳，该时间戳与提交文件的时间戳可能存在几秒的误差
                    # 因此，我们只根据企业名称和报告类型来查找，并选择最新的一个
                    
                    found_reports = []
                    for f in os.listdir(reports_folder):
                        if enterprise_name in f and f.endswith(('.docx', '.pdf')):
                            filepath = os.path.join(reports_folder, f)
                            found_reports.append((os.path.getmtime(filepath), f))
                    
                    if found_reports:
                        # 按修改时间降序排序，取最新的一个
                        found_reports.sort(key=lambda x: x[0], reverse=True)
                        report_filename = found_reports[0][1]

                history_items.append({
                    'time_text': datetime.fromtimestamp(sub['submit_time']).strftime('%Y-%m-%d %H:%M'),
                    'user_level': submission_data.get('user_level', '未知'),
                    'score': score_info,
                    'excel_file': os.path.basename(excel_path),
                    'report_file': report_filename
                })
            except Exception:
                continue

        return jsonify({'success': True, 'items': history_items})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/enterprise/reports')
@role_required('enterprise')
def api_enterprise_reports():
    """获取当前企业的所有报告"""
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'error': '无法获取用户信息，请重新登录'}), 401

        # 获取企业名称
        display_name = session.get('display_name', username)

        reports_folder = app.config['REPORT_FOLDER']
        reports = []

        if os.path.exists(reports_folder):
            for filename in os.listdir(reports_folder):
                if display_name in filename and filename.endswith(('.docx', '.pdf')):
                    filepath = os.path.join(reports_folder, filename)
                    stat = os.stat(filepath)
                    reports.append({
                        'filename': filename,
                        'type': 'Word' if filename.endswith('.docx') else 'PDF',
                        'file_size': stat.st_size,
                        'time_text': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M')
                    })

        reports.sort(key=lambda x: x['time_text'], reverse=True)

        return jsonify({'success': True, 'items': reports})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/portal/chamber/send-report', methods=['POST'])
@role_required('chamber_of_commerce')
def api_chamber_send_report():
    """手动发送报告邮件"""
    try:
        data = request.get_json()
        recipient = data.get('recipient')
        report_filename = data.get('report_filename')

        if not recipient or not report_filename:
            return jsonify({'success': False, 'error': '参数不完整'}), 400

        report_path = os.path.join(app.config['REPORT_FOLDER'], report_filename)
        if not os.path.exists(report_path):
            return jsonify({'success': False, 'error': '报告文件不存在'}), 404

        success = notification_service.send_email(
            to_email=recipient,
            enterprise_name='企业',
            contact_name='您好',
            report_url='',
            attachment_path=report_path
        )

        if success:
            return jsonify({'success': True, 'message': '邮件发送成功'})
        else:
            return jsonify({'success': False, 'error': '邮件发送失败'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/portal/chamber/tutoring-records')
@role_required('chamber_of_commerce')
def api_chamber_tutoring_records():
    """获取专家辅导记录"""
    # 这是一个占位符。在实际应用中，您将从数据库中获取这些数据。
    # 现在，我们模拟一些数据以便预览。
    records = [
        {'expert': 'expert_A', 'enterprise': '企业示例', 'message': '建议关注一下公司治理结构。', 'time': '2025-11-26 10:00'},
        {'expert': 'expert_B', 'enterprise': '另一家公司', 'message': '科技创新方面有待加强。', 'time': '2025-11-25 15:30'},
    ]
    return jsonify({'success': True, 'records': records})


if __name__ == '__main__':
    print("=" * 60)
    print("企业现代制度评价系统启动")
    print("=" * 60)
    print("功能：")
    print("  1. 客户端首页: http://localhost:5000")
    print("  2. 在线问卷填写: http://localhost:5000/questionnaire")
    print("  3. 管理员登录: http://localhost:5000/admin/login")
    print("     默认账号: admin")
    print("     默认密码: admin123")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)
