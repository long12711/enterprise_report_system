"""
问卷生成器 - 基于新的指标体系
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os


class QuestionnaireGenerator:
    """问卷生成器类"""

    def __init__(self, indicator_file=None):
        """
        初始化问卷生成器

        Args:
            indicator_file: 指标体系Excel文件路径
        """
        if indicator_file is None:
            indicator_file = r'D:\Claude Code\enterprise_report_system\指标体系.xlsx'

        self.indicator_file = indicator_file
        self.indicators_df = None
        self.load_indicators()

    def load_indicators(self):
        """加载指标体系"""
        try:
            self.indicators_df = pd.read_excel(self.indicator_file, sheet_name=0)
            print(f"[OK] 成功加载指标体系: {len(self.indicators_df)} 个问题")
        except Exception as e:
            print(f"[ERROR] 加载指标体系失败: {e}")
            raise

    def generate_questionnaire(self, output_path=None, enterprise_name=None):
        """
        生成问卷Excel文件

        Args:
            output_path: 输出文件路径
            enterprise_name: 企业名称（可选，用于定制化问卷）

        Returns:
            生成的文件路径
        """
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'问卷_{timestamp}.xlsx'

        print(f"\n开始生成问卷: {output_path}")

        # 创建工作簿
        wb = Workbook()

        # 删除默认的Sheet
        wb.remove(wb.active)

        # 1. 创建"企业信息"工作表
        self._create_enterprise_info_sheet(wb, enterprise_name)

        # 2. 创建"问卷填写说明"工作表
        self._create_instruction_sheet(wb)

        # 3. 创建"问卷"工作表
        self._create_questionnaire_sheet(wb)

        # 4. 创建"指标说明"工作表
        self._create_indicator_guide_sheet(wb)

        # 保存文件
        wb.save(output_path)
        print(f"[OK] 问卷生成成功: {output_path}")

        return output_path

    def _create_enterprise_info_sheet(self, wb, enterprise_name=None):
        """创建企业信息工作表"""
        ws = wb.create_sheet('企业信息', 0)

        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

        # 标题样式
        title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 边框样式
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 添加标题
        ws['A1'] = '现代企业制度指数评价 - 企业信息表'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 30

        # 企业基本信息
        info_fields = [
            ('企业名称', enterprise_name or ''),
            ('统一社会信用代码', ''),
            ('企业类型', '请选择：有限责任公司/股份有限公司/个人独资企业/合伙企业/其他'),
            ('成立时间', 'YYYY-MM-DD'),
            ('注册资本（万元）', ''),
            ('所属行业', ''),
            ('员工人数', ''),
            ('年营业收入（万元）', ''),
            ('联系人姓名', ''),
            ('联系人职务', ''),
            ('联系人手机', ''),
            ('联系人邮箱', ''),
            ('企业地址', ''),
            ('填写日期', datetime.now().strftime('%Y-%m-%d')),
        ]

        row = 3
        for field_name, field_value in info_fields:
            ws[f'A{row}'] = field_name
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].border = border
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

            ws[f'B{row}'] = field_value
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

            ws.row_dimensions[row].height = 25
            row += 1

        # 添加填写说明
        row += 1
        ws[f'A{row}'] = '填写说明：'
        ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        instructions = [
            '1. 请如实填写企业基本信息，带*为必填项',
            '2. 企业类型请根据营业执照准确选择',
            '3. 联系人信息用于接收评价报告，请确保准确无误',
            '4. 填写完成后，请继续填写"问卷"工作表'
        ]
        for instruction in instructions:
            ws[f'A{row}'] = instruction
            ws[f'A{row}'].font = Font(name='微软雅黑', size=9, color='666666')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

    def _create_instruction_sheet(self, wb):
        """创建填写说明工作表"""
        ws = wb.create_sheet('问卷填写说明', 1)

        # 设置列宽
        ws.column_dimensions['A'].width = 100

        # 标题
        ws['A1'] = '现代企业制度指数评价问卷 - 填写说明'
        ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='4472C4')
        ws.row_dimensions[1].height = 35

        # 说明内容
        instructions = [
            '',
            '一、问卷概述',
            '本问卷旨在全面评估企业现代制度建设情况，包括9大一级指标、多个二级和三级指标，共231个评价问题。',
            '',
            '二、指标体系',
            '1. 党建引领：评估党组织建设、政治引领、党建参与治理等方面',
            '2. 产权结构：评估权属清晰、法人财产独立、股权结构合理性',
            '3. 公司治理结构和机制：评估股东会、董事会、监事会等治理机制',
            '4. 战略管理：评估战略规划、导向和执行情况',
            '5. 内控、风险与合规管理：评估内部控制、风险管理和合规体系',
            '6. 科学民主管理：评估科学管理和员工权益保障',
            '7. 科技创新：评估创新战略和成果转化能力',
            '8. 社会责任与企业文化：评估社会责任履行和文化建设',
            '9. 家族企业治理：针对家族企业的特殊治理问题',
            '',
            '三、填写方法',
            '1. 答题选项说明：',
            '   - 合规类：选择"是"或"否"',
            '   - 有效性：选择"很有效"、"比较有效"、"一般"、"不太有效"、"完全无效"',
            '   - 一票否决：如选择"否"，该项直接扣分',
            '',
            '2. 计分规则：',
            '   - 合规类问题："是"得满分，"否"得0分',
            '   - 有效性问题：根据选择程度按比例得分',
            '   - 一票否决项：选择"否"直接扣除相应分值（负分）',
            '',
            '3. 适用对象：',
            '   - 所有企业：适用于所有参评企业',
            '   - 公司制企业：仅适用于有限责任公司和股份有限公司',
            '   - 股份有限公司：仅适用于股份有限公司',
            '   - 有限责任公司：仅适用于有限责任公司',
            '   如问题不适用您的企业类型，可填写"不适用"',
            '',
            '四、注意事项',
            '1. 请根据企业实际情况如实填写，确保信息准确性',
            '2. 对于不清楚的问题，建议咨询相关部门负责人后填写',
            '3. 填写过程中可参考"指标说明"工作表中的详细解释',
            '4. 建议由企业管理层组织相关部门协同完成',
            '5. 填写完成后请检查是否有遗漏项',
            '',
            '五、联系方式',
            '如有疑问，请联系：',
            '联系人：[待填写]',
            '电话：[待填写]',
            '邮箱：[待填写]',
        ]

        row = 2
        for instruction in instructions:
            ws[f'A{row}'] = instruction
            if instruction.startswith('一、') or instruction.startswith('二、') or instruction.startswith('三、') or instruction.startswith('四、') or instruction.startswith('五、'):
                ws[f'A{row}'].font = Font(name='微软雅黑', size=12, bold=True, color='4472C4')
                ws.row_dimensions[row].height = 25
            elif instruction.startswith(('1. ', '2. ', '3. ', '4. ', '5. ', '6. ', '7. ', '8. ', '9. ')):
                ws[f'A{row}'].font = Font(name='微软雅黑', size=10, bold=True)
            else:
                ws[f'A{row}'].font = Font(name='微软雅黑', size=10)
            row += 1

    def _create_questionnaire_sheet(self, wb):
        """创建问卷工作表"""
        ws = wb.create_sheet('问卷', 2)

        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 15  # 一级指标
        ws.column_dimensions['C'].width = 18  # 二级指标
        ws.column_dimensions['D'].width = 60  # 三级指标（问题）
        ws.column_dimensions['E'].width = 15  # 指标类型
        ws.column_dimensions['F'].width = 8   # 分值
        ws.column_dimensions['G'].width = 20  # 适用对象
        ws.column_dimensions['H'].width = 30  # 答案（下拉选择）
        ws.column_dimensions['I'].width = 40  # 备注说明

        # 样式定义
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 创建表头
        headers = ['序号', '一级指标', '二级指标', '三级指标（问题）', '指标类型', '分值', '适用对象', '答案', '备注说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[1].height = 30

        # 填充问题数据
        row = 2
        current_level1 = None
        current_level2 = None

        for idx, indicator_row in self.indicators_df.iterrows():
            seq_no = indicator_row['序号']
            level1 = indicator_row['一级指标']
            level2 = indicator_row['二级指标']
            level3 = indicator_row['三级指标']
            indicator_type = indicator_row['指标类型']
            score = indicator_row['分值']
            applicable = indicator_row['适用对象']

            # 处理一级指标（合并单元格）
            if pd.notna(level1):
                current_level1 = level1

            # 处理二级指标
            if pd.notna(level2):
                current_level2 = level2

            # 填充数据
            ws.cell(row=row, column=1, value=int(seq_no) if pd.notna(seq_no) else '')
            ws.cell(row=row, column=2, value=current_level1 if current_level1 else '')
            ws.cell(row=row, column=3, value=current_level2 if current_level2 else '')
            ws.cell(row=row, column=4, value=level3 if pd.notna(level3) else '')
            ws.cell(row=row, column=5, value=indicator_type if pd.notna(indicator_type) else '')
            ws.cell(row=row, column=6, value=score if pd.notna(score) else '')
            ws.cell(row=row, column=7, value=applicable if pd.notna(applicable) else '所有企业')
            ws.cell(row=row, column=8, value='')  # 答案列，留空待填写
            ws.cell(row=row, column=9, value='')  # 备注列

            # 设置样式
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 设置答案列的数据验证（下拉选择）
            answer_cell = ws.cell(row=row, column=8)
            if pd.notna(indicator_type):
                if '合规' in str(indicator_type):
                    from openpyxl.worksheet.datavalidation import DataValidation
                    dv = DataValidation(type="list", formula1='"是,否,不适用"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(answer_cell)
                elif '有效' in str(indicator_type):
                    dv = DataValidation(type="list", formula1='"很有效,比较有效,一般,不太有效,完全无效,不适用"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(answer_cell)
                elif '一票否决' in str(indicator_type) or '否决' in str(indicator_type):
                    dv = DataValidation(type="list", formula1='"是,否,不适用"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(answer_cell)
                    # 一票否决项标红
                    ws.cell(row=row, column=5).font = Font(color='FF0000', bold=True)

            ws.row_dimensions[row].height = 40
            row += 1

        # 冻结首行
        ws.freeze_panes = 'A2'

    def _create_indicator_guide_sheet(self, wb):
        """创建指标说明工作表"""
        ws = wb.create_sheet('指标说明', 3)

        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 15

        # 样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ['一级指标', '二级指标', '指标说明', '对应指引条目']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[1].height = 25

        # 填充指标说明
        row = 2
        current_level1 = None

        # 提取唯一的二级指标及其说明
        for idx, indicator_row in self.indicators_df.iterrows():
            level1 = indicator_row['一级指标']
            level2 = indicator_row['二级指标']
            guide_ref = indicator_row['对应指引条目']

            if pd.notna(level1):
                current_level1 = level1

            if pd.notna(level2):
                ws.cell(row=row, column=1, value=current_level1)
                ws.cell(row=row, column=2, value=level2)
                ws.cell(row=row, column=3, value=f'{level2}相关问题的详细说明和评价标准')
                ws.cell(row=row, column=4, value=guide_ref if pd.notna(guide_ref) else '')

                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                row += 1

        ws.freeze_panes = 'A2'

    def generate_batch_questionnaires(self, enterprises_df, output_folder='questionnaires'):
        """
        批量生成问卷

        Args:
            enterprises_df: 包含企业信息的DataFrame
            output_folder: 输出文件夹

        Returns:
            生成的文件列表
        """
        os.makedirs(output_folder, exist_ok=True)
        generated_files = []

        for idx, row in enterprises_df.iterrows():
            enterprise_name = row.get('企业名称', f'企业{idx+1}')
            output_path = os.path.join(
                output_folder,
                f'问卷_{enterprise_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )

            try:
                file_path = self.generate_questionnaire(output_path, enterprise_name)
                generated_files.append(file_path)
                print(f"[OK] 已生成: {enterprise_name}")
            except Exception as e:
                print(f"[ERROR] 生成失败 {enterprise_name}: {e}")

        return generated_files


if __name__ == '__main__':
    # 测试问卷生成
    generator = QuestionnaireGenerator()

    # 生成单个问卷
    output_file = generator.generate_questionnaire(
        output_path='测试问卷.xlsx',
        enterprise_name='测试科技有限公司'
    )

    print(f"\n问卷已生成: {output_file}")
