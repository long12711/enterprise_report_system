# -*- coding: utf-8 -*-
"""
工商联用户管理 - 完整实现
包含：数据库操作、权限检查、API 接口、日志记录
"""

from flask import Blueprint, request, jsonify, session
from datetime import datetime
import uuid
import json
import logging
from functools import wraps
import bcrypt

# 配置日志
logger = logging.getLogger(__name__)

# 创建蓝图
chamber_users_bp = Blueprint('chamber_users', __name__, url_prefix='/api/portal/chamber')

# ============================================================================
# 数据库连接（使用 Flask-SQLAlchemy）
# ============================================================================

def get_db():
    """获取数据库连接"""
    try:
        from app import db
        return db
    except:
        return None


# ============================================================================
# 权限检查工具类
# ============================================================================

class PermissionChecker:
    """权限检查工具"""
    
    @staticmethod
    def get_current_user():
        """获取当前登录用户"""
        user_id = session.get('user_id')
        if not user_id:
            return None
        
        try:
            db = get_db()
            if not db:
                return None
            
            # 使用原生 SQL 查询
            from sqlalchemy import text
            result = db.session.execute(
                text('SELECT * FROM chamber_users WHERE id = :id'),
                {'id': user_id}
            ).fetchone()
            
            if result:
                return dict(result._mapping) if hasattr(result, '_mapping') else dict(result)
            return None
        except Exception as e:
            logger.error(f"获取当前用户失败: {e}")
            return None
    
    @staticmethod
    def can_view_user(current_user, target_user):
        """检查是否可以查看用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以看所有用户
        if current_user.get('level') == 'national':
            return True
        
        # 省级管理员可以看本省的用户
        if current_user.get('level') == 'province':
            return target_user.get('level') in ['county', 'province'] and \
                   target_user.get('region') == current_user.get('region')
        
        # 县市级管理员只能看本县市的用户
        if current_user.get('level') == 'county':
            return target_user.get('level') == 'county' and \
                   target_user.get('region') == current_user.get('region')
        
        return False
    
    @staticmethod
    def can_edit_user(current_user, target_user):
        """检查是否可以编辑用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以编辑所有用户
        if current_user.get('level') == 'national':
            return True
        
        # 省级管理员可以编辑本省的非管理员用户
        if current_user.get('level') == 'province':
            if target_user.get('role') == 'admin':
                return False
            return target_user.get('region') == current_user.get('region') and \
                   target_user.get('level') in ['county', 'province']
        
        # 县市级管理员只能编辑本县市的操作员
        if current_user.get('level') == 'county':
            if target_user.get('role') in ['admin', 'reviewer']:
                return False
            return target_user.get('region') == current_user.get('region') and \
                   target_user.get('level') == 'county'
        
        return False
    
    @staticmethod
    def can_delete_user(current_user, target_user):
        """检查是否可以删除用户"""
        if not current_user or not target_user:
            return False
        
        # 全联管理员可以删除所有用户
        if current_user.get('level') == 'national':
            return True
        
        # 省级管理员可以删除本省的操作员
        if current_user.get('level') == 'province':
            if target_user.get('role') != 'operator':
                return False
            return target_user.get('region') == current_user.get('region') and \
                   target_user.get('level') in ['county', 'province']
        
        # 县市级管理员只能删除本县市的操作员
        if current_user.get('level') == 'county':
            if target_user.get('role') != 'operator':
                return False
            return target_user.get('region') == current_user.get('region') and \
                   target_user.get('level') == 'county'
        
        return False
    
    @staticmethod
    def can_create_user(current_user, new_user_level, new_user_region):
        """检查是否可以创建用户"""
        if not current_user:
            return False
        
        # 权限层级
        level_hierarchy = {'county': 1, 'province': 2, 'national': 3}
        
        # 不能创建高于自己权限的用户
        if level_hierarchy.get(new_user_level, 0) > level_hierarchy.get(current_user.get('level'), 0):
            return False
        
        # 检查地区权限
        if current_user.get('level') == 'national':
            return True
        
        if current_user.get('level') == 'province':
            return new_user_level in ['county', 'province'] and \
                   new_user_region == current_user.get('region')
        
        if current_user.get('level') == 'county':
            return new_user_level == 'county' and \
                   new_user_region == current_user.get('region')
        
        return False


# ============================================================================
# 日志记录函数
# ============================================================================

def log_operation(operator_id, target_user_id, action, old_value=None, new_value=None):
    """记录用户操作日志"""
    try:
        db = get_db()
        if not db:
            return False
        
        from sqlalchemy import text
        
        log_id = str(uuid.uuid4())
        
        db.session.execute(
            text('''
                INSERT INTO chamber_user_logs 
                (id, operator_id, target_user_id, action, old_value, new_value)
                VALUES (:id, :operator_id, :target_user_id, :action, :old_value, :new_value)
            '''),
            {
                'id': log_id,
                'operator_id': operator_id,
                'target_user_id': target_user_id,
                'action': action,
                'old_value': json.dumps(old_value, ensure_ascii=False) if old_value else None,
                'new_value': json.dumps(new_value, ensure_ascii=False) if new_value else None
            }
        )
        db.session.commit()
        return True
    except Exception as e:
        logger.error(f"记录操作日志失败: {e}")
        return False


# ============================================================================
# 装饰器：权限检查
# ============================================================================

def require_login(f):
    """要求登录"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        current_user = PermissionChecker.get_current_user()
        if not current_user:
            return jsonify({'code': 401, 'message': '未登录'}), 401
        return f(*args, **kwargs)
    return decorated_function


def require_chamber_admin(f):
    """要求工商联管理员权限"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        current_user = PermissionChecker.get_current_user()
        if not current_user or current_user.get('role') not in ['admin', 'reviewer']:
            return jsonify({'code': 403, 'message': '权限不足'}), 403
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# API 接口
# ============================================================================

@chamber_users_bp.route('/users', methods=['GET'])
@require_login
@require_chamber_admin
def list_users():
    """获取用户列表（支持分页和筛选）"""
    try:
        current_user = PermissionChecker.get_current_user()
        
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 10, type=int)
        level = request.args.get('level', '')
        region = request.args.get('region', '')
        role = request.args.get('role', '')
        status = request.args.get('status', '')
        keyword = request.args.get('keyword', '')
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 构建查询条件
        conditions = []
        params = {}
        
        # 权限检查：只能看到自己权限范围内的用户
        if current_user.get('level') == 'national':
            pass  # 全联可以看所有
        elif current_user.get('level') == 'province':
            conditions.append('(level IN ("county", "province") AND region = :region)')
            params['region'] = current_user.get('region')
        elif current_user.get('level') == 'county':
            conditions.append('(level = "county" AND region = :region)')
            params['region'] = current_user.get('region')
        
        # 其他筛选条件
        if level:
            conditions.append('level = :level')
            params['level'] = level
        
        if region:
            conditions.append('region = :region')
            params['region'] = region
        
        if role:
            conditions.append('role = :role')
            params['role'] = role
        
        if status:
            conditions.append('status = :status')
            params['status'] = status
        
        if keyword:
            conditions.append('(username LIKE :keyword OR real_name LIKE :keyword OR email LIKE :keyword)')
            params['keyword'] = f'%{keyword}%'
        
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        # 获取总数
        count_sql = f'SELECT COUNT(*) as total FROM chamber_users WHERE {where_clause}'
        count_result = db.session.execute(text(count_sql), params).fetchone()
        total = dict(count_result._mapping)['total'] if count_result else 0
        
        # 获取分页数据
        offset = (page - 1) * page_size
        data_sql = f'''
            SELECT id, username, email, real_name, phone, level, region, role, 
                   review_level, department, position, status, created_at, updated_at
            FROM chamber_users 
            WHERE {where_clause}
            ORDER BY created_at DESC
            LIMIT :offset, :limit
        '''
        params['offset'] = offset
        params['limit'] = page_size
        
        results = db.session.execute(text(data_sql), params).fetchall()
        users = [dict(row._mapping) for row in results]
        
        # 转换时间格式
        for user in users:
            if user.get('created_at'):
                user['created_at'] = user['created_at'].isoformat()
            if user.get('updated_at'):
                user['updated_at'] = user['updated_at'].isoformat()
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'users': users,
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size
            }
        })
    
    except Exception as e:
        logger.error(f"获取用户列表失败: {e}")
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['GET'])
@require_login
def get_user(user_id):
    """获取单个用户详情"""
    try:
        current_user = PermissionChecker.get_current_user()
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 获取目标用户
        result = db.session.execute(
            text('SELECT * FROM chamber_users WHERE id = :id'),
            {'id': user_id}
        ).fetchone()
        
        if not result:
            return jsonify({'code': 404, 'message': '用户不存在'}), 404
        
        target_user = dict(result._mapping)
        
        # 权限检查
        if not PermissionChecker.can_view_user(current_user, target_user):
            return jsonify({'code': 403, 'message': '权限不足'}), 403
        
        # 转换时间格式
        if target_user.get('created_at'):
            target_user['created_at'] = target_user['created_at'].isoformat()
        if target_user.get('updated_at'):
            target_user['updated_at'] = target_user['updated_at'].isoformat()
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': target_user
        })
    
    except Exception as e:
        logger.error(f"获取用户详情失败: {e}")
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


@chamber_users_bp.route('/users', methods=['POST'])
@require_login
@require_chamber_admin
def create_user():
    """创建新用户"""
    try:
        current_user = PermissionChecker.get_current_user()
        data = request.get_json()
        
        # 验证必填字段
        required_fields = ['username', 'email', 'password', 'real_name', 'level', 'region', 'role']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400
        
        # 权限检查
        if not PermissionChecker.can_create_user(current_user, data['level'], data['region']):
            return jsonify({'code': 403, 'message': '权限不足，无法创建此级别的用户'}), 403
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 检查用户名和邮箱是否已存在
        existing = db.session.execute(
            text('SELECT id FROM chamber_users WHERE username = :username OR email = :email'),
            {'username': data['username'], 'email': data['email']}
        ).fetchone()
        
        if existing:
            return jsonify({'code': 400, 'message': '用户名或邮箱已存在'}), 400
        
        # 密码加密
        password_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # 创建用户
        user_id = str(uuid.uuid4())
        
        db.session.execute(
            text('''
                INSERT INTO chamber_users 
                (id, username, email, password, real_name, phone, level, region, role, 
                 review_level, department, position, status, remark, created_by)
                VALUES (:id, :username, :email, :password, :real_name, :phone, :level, :region, :role,
                        :review_level, :department, :position, :status, :remark, :created_by)
            '''),
            {
                'id': user_id,
                'username': data['username'],
                'email': data['email'],
                'password': password_hash,
                'real_name': data['real_name'],
                'phone': data.get('phone'),
                'level': data['level'],
                'region': data['region'],
                'role': data['role'],
                'review_level': data.get('review_level'),
                'department': data.get('department'),
                'position': data.get('position'),
                'status': data.get('status', 'pending'),
                'remark': data.get('remark'),
                'created_by': current_user['id']
            }
        )
        db.session.commit()
        
        # 记录日志
        log_operation(current_user['id'], user_id, 'create', None, data)
        
        return jsonify({
            'code': 201,
            'message': '创建成功',
            'data': {'id': user_id}
        }), 201
    
    except Exception as e:
        logger.error(f"创建用户失败: {e}")
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['PUT'])
@require_login
@require_chamber_admin
def update_user(user_id):
    """更新用户信息"""
    try:
        current_user = PermissionChecker.get_current_user()
        data = request.get_json()
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 获取目标用户
        result = db.session.execute(
            text('SELECT * FROM chamber_users WHERE id = :id'),
            {'id': user_id}
        ).fetchone()
        
        if not result:
            return jsonify({'code': 404, 'message': '用户不存在'}), 404
        
        target_user = dict(result._mapping)
        
        # 权限检查
        if not PermissionChecker.can_edit_user(current_user, target_user):
            return jsonify({'code': 403, 'message': '权限不足'}), 403
        
        # 构建更新语句
        update_fields = []
        params = {'id': user_id}
        
        allowed_fields = ['real_name', 'phone', 'role', 'review_level', 'department', 'position', 'status', 'remark']
        
        for field in allowed_fields:
            if field in data:
                update_fields.append(f'{field} = :{field}')
                params[field] = data[field]
        
        if not update_fields:
            return jsonify({'code': 400, 'message': '没有可更新的字段'}), 400
        
        # 执行更新
        update_sql = f'''
            UPDATE chamber_users 
            SET {', '.join(update_fields)}, updated_at = NOW()
            WHERE id = :id
        '''
        
        db.session.execute(text(update_sql), params)
        db.session.commit()
        
        # 记录日志
        old_value = {k: target_user.get(k) for k in allowed_fields}
        new_value = {k: data.get(k) for k in allowed_fields if k in data}
        log_operation(current_user['id'], user_id, 'update', old_value, new_value)
        
        return jsonify({
            'code': 200,
            'message': '更新成功'
        })
    
    except Exception as e:
        logger.error(f"更新用户失败: {e}")
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500


@chamber_users_bp.route('/users/<user_id>', methods=['DELETE'])
@require_login
@require_chamber_admin
def delete_user(user_id):
    """删除用户"""
    try:
        current_user = PermissionChecker.get_current_user()
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 获取目标用户
        result = db.session.execute(
            text('SELECT * FROM chamber_users WHERE id = :id'),
            {'id': user_id}
        ).fetchone()
        
        if not result:
            return jsonify({'code': 404, 'message': '用户不存在'}), 404
        
        target_user = dict(result._mapping)
        
        # 权限检查
        if not PermissionChecker.can_delete_user(current_user, target_user):
            return jsonify({'code': 403, 'message': '权限不足'}), 403
        
        # 删除用户
        db.session.execute(
            text('DELETE FROM chamber_users WHERE id = :id'),
            {'id': user_id}
        )
        db.session.commit()
        
        # 记录日志
        log_operation(current_user['id'], user_id, 'delete', target_user, None)
        
        return jsonify({
            'code': 200,
            'message': '删除成功'
        })
    
    except Exception as e:
        logger.error(f"删除用户失败: {e}")
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500


@chamber_users_bp.route('/users/export', methods=['GET'])
@require_login
@require_chamber_admin
def export_users():
    """导出用户列表为 Excel"""
    try:
        current_user = PermissionChecker.get_current_user()
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # 构建查询条件（同列表接口）
        conditions = []
        params = {}
        
        if current_user.get('level') == 'national':
            pass
        elif current_user.get('level') == 'province':
            conditions.append('(level IN ("county", "province") AND region = :region)')
            params['region'] = current_user.get('region')
        elif current_user.get('level') == 'county':
            conditions.append('(level = "county" AND region = :region)')
            params['region'] = current_user.get('region')
        
        where_clause = ' AND '.join(conditions) if conditions else '1=1'
        
        # 获取数据
        sql = f'''
            SELECT id, username, email, real_name, phone, level, region, role, 
                   review_level, department, position, status, created_at
            FROM chamber_users 
            WHERE {where_clause}
            ORDER BY created_at DESC
        '''
        
        results = db.session.execute(text(sql), params).fetchall()
        
        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '工商联用户'
        
        # 设置表头
        headers = ['用户ID', '用户名', '邮箱', '真实姓名', '手机号', '层级', '地区', '角色', 
                   '审核权限', '部门', '职位', '状态', '创建时间']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加数据
        for row in results:
            row_dict = dict(row._mapping)
            ws.append([
                row_dict['id'],
                row_dict['username'],
                row_dict['email'],
                row_dict['real_name'],
                row_dict['phone'],
                row_dict['level'],
                row_dict['region'],
                row_dict['role'],
                row_dict['review_level'],
                row_dict['department'],
                row_dict['position'],
                row_dict['status'],
                row_dict['created_at'].isoformat() if row_dict['created_at'] else ''
            ])
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'工商联用户_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        logger.error(f"导出用户失败: {e}")
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


@chamber_users_bp.route('/logs', methods=['GET'])
@require_login
@require_chamber_admin
def get_logs():
    """获取操作日志"""
    try:
        current_user = PermissionChecker.get_current_user()
        
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 10, type=int)
        
        db = get_db()
        if not db:
            return jsonify({'code': 500, 'message': '数据库连接失败'}), 500
        
        from sqlalchemy import text
        
        # 获取总数
        count_result = db.session.execute(
            text('SELECT COUNT(*) as total FROM chamber_user_logs')
        ).fetchone()
        total = dict(count_result._mapping)['total'] if count_result else 0
        
        # 获取分页数据
        offset = (page - 1) * page_size
        results = db.session.execute(
            text('''
                SELECT id, operator_id, target_user_id, action, old_value, new_value, created_at
                FROM chamber_user_logs
                ORDER BY created_at DESC
                LIMIT :offset, :limit
            '''),
            {'offset': offset, 'limit': page_size}
        ).fetchall()
        
        logs = []
        for row in results:
            log_dict = dict(row._mapping)
            log_dict['created_at'] = log_dict['created_at'].isoformat() if log_dict['created_at'] else None
            if log_dict.get('old_value'):
                log_dict['old_value'] = json.loads(log_dict['old_value'])
            if log_dict.get('new_value'):
                log_dict['new_value'] = json.loads(log_dict['new_value'])
            logs.append(log_dict)
        
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'logs': logs,
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size
            }
        })
    
    except Exception as e:
        logger.error(f"获取操作日志失败: {e}")
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500

