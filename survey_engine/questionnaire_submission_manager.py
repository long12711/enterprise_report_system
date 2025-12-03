"""
问卷提交管理模块
功能：接收在线问卷提交、保存数据、生成Excel文件
"""
import os
import json
from datetime import datetime
import pandas as pd
from questionnaire_generator import QuestionnaireGenerator
from score_calculator import ScoreCalculator


class QuestionnaireSubmissionManager:
    """问卷提交管理器"""

    def __init__(self, storage_folder='submissions'):
        """
        初始化管理器

        Args:
            storage_folder: 提交数据存储文件夹
        """
        self.storage_folder = storage_folder
        self.calculator = ScoreCalculator()
        os.makedirs(storage_folder, exist_ok=True)

    def save_submission(self, submission_data):
        """
        保存问卷提交数据

        Args:
            submission_data: 包含企业信息和答案的字典
                {
                    'enterprise_info': {...},
                    'answers': {question_sequence: answer, ...},
                    'scores': {question_sequence: score, ...},
                    'questions_snapshot': [ {sequence, level1, level2, question, question_type, base_score, criteria...}, ...],
                    'user_type': 'enterprise',
                    'user_level': 'beginner'
                }

        Returns:
            保存的文件路径
        """
        try:
            # 生成文件名
            enterprise_name = submission_data['enterprise_info'].get('企业名称', '未命名企业')
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            # 保存为JSON（备份）
            json_filename = f"submission_{enterprise_name}_{timestamp}.json"
            json_filepath = os.path.join(self.storage_folder, json_filename)

            with open(json_filepath, 'w', encoding='utf-8') as f:
                json.dump(submission_data, f, ensure_ascii=False, indent=2)

            # 生成Excel文件（用于后续处理）
            excel_filename = f"问卷_{enterprise_name}_{timestamp}.xlsx"
            excel_filepath = os.path.join(self.storage_folder, excel_filename)

            self._create_filled_questionnaire(submission_data, excel_filepath)

            return {
                'json_path': json_filepath,
                'excel_path': excel_filepath,
                'enterprise_name': enterprise_name,
                'timestamp': timestamp
            }

        except Exception as e:
            print(f"[ERROR] 保存提交数据失败: {e}")
            raise

    def _create_filled_questionnaire(self, submission_data, output_path):
        """
        创建已填写的问卷Excel文件

        Args:
            submission_data: 提交数据
            output_path: 输出文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows

            # 题目来源：优先使用提交时的题目快照（与本次分级一致）；否则退回读取本地指标文件
            questions_snapshot = submission_data.get('questions_snapshot')

            # 创建工作簿
            wb = Workbook()

            # 1. 企业信息工作表
            ws_info = wb.active
            ws_info.title = "企业信息"

            enterprise_info = submission_data['enterprise_info']
            info_data = [
                ['项目', '内容'],
                ['企业名称', enterprise_info.get('企业名称', '')],
                ['统一社会信用代码', enterprise_info.get('统一社会信用代码', '')],
                ['企业类型', enterprise_info.get('企业类型', '')],
                ['所属行业', enterprise_info.get('所属行业', '')],
                ['注册资本（万元）', enterprise_info.get('注册资本（万元）', '')],
                ['成立时间', enterprise_info.get('成立时间', '')],
                ['员工人数', enterprise_info.get('员工人数', '')],
                ['年营业收入（万元）', enterprise_info.get('年营业收入（万元）', '')],
                ['联系人姓名', enterprise_info.get('联系人姓名', '')],
                ['联系人邮箱', enterprise_info.get('联系人邮箱', '')],
                ['联系人电话', enterprise_info.get('联系人电话', '')],
                ['填写时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]

            for row in info_data:
                ws_info.append(row)

            # 设置样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)

            ws_info['A1'].fill = header_fill
            ws_info['B1'].fill = header_fill
            ws_info['A1'].font = header_font
            ws_info['B1'].font = header_font

            ws_info.column_dimensions['A'].width = 25
            ws_info.column_dimensions['B'].width = 40

            # 2. 问卷工作表
            ws_questionnaire = wb.create_sheet("问卷")

            # 表头（兼容清单打分）
            headers = ['序号', '一级指标', '二级指标', '三级指标（问题）', '问题类型', '分值', '答案/清单选择', '打分标准', '计算分数']
            ws_questionnaire.append(headers)

            # 设置表头样式
            for col_num, header in enumerate(headers, 1):
                cell = ws_questionnaire.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 填充问题和答案
            answers = submission_data.get('answers', {})
            scores = submission_data.get('scores', {})

            if questions_snapshot:
                for q in questions_snapshot:
                    seq = q.get('sequence')
                    seq_str = str(seq)
                    answer_raw = answers.get(seq_str, '')
                    # 如为JSON字符串（清单题），直接写入，便于后续解析
                    try:
                        # 尝试缩短显示
                        disp = answer_raw
                        if isinstance(answer_raw, str) and answer_raw.startswith('{'):
                            disp = answer_raw
                        answer_disp = disp
                    except Exception:
                        answer_disp = answer_raw

                    data_row = [
                        seq,
                        q.get('level1', ''),
                        q.get('level2', ''),
                        q.get('question', ''),
                        q.get('question_type', ''),
                        q.get('base_score', ''),
                        answer_disp,
                        q.get('criteria', ''),
                        scores.get(seq_str, '')
                    ]
                    ws_questionnaire.append(data_row)
            else:
                # 兼容旧路径：读取本地指标文件
                indicators_df = pd.read_excel('指标体系.xlsx')
                for idx, row in indicators_df.iterrows():
                    sequence = row.get('序号', idx + 1)
                    seq_str = str(sequence)
                    answer = answers.get(seq_str, '未填写')
                    data_row = [
                        sequence,
                        row.get('一级指标', ''),
                        row.get('二级指标', ''),
                        row.get('三级指标（问题）', ''),
                        row.get('问题类型', ''),
                        row.get('分值', 0),
                        answer,
                        '',
                        scores.get(seq_str, '')
                    ]
                    ws_questionnaire.append(data_row)

            # 设置列宽
            ws_questionnaire.column_dimensions['A'].width = 8
            ws_questionnaire.column_dimensions['B'].width = 20
            ws_questionnaire.column_dimensions['C'].width = 20
            ws_questionnaire.column_dimensions['D'].width = 52
            ws_questionnaire.column_dimensions['E'].width = 12
            ws_questionnaire.column_dimensions['F'].width = 8
            ws_questionnaire.column_dimensions['G'].width = 26
            ws_questionnaire.column_dimensions['H'].width = 46
            ws_questionnaire.column_dimensions['I'].width = 10

            # 保存文件
            wb.save(output_path)
            print(f"[OK] 已填写问卷文件已生成: {output_path}")

        except Exception as e:
            print(f"[ERROR] 生成Excel文件失败: {e}")
            raise

    def get_all_submissions(self):
        """
        获取所有提交记录

        Returns:
            提交记录列表
        """
        submissions = []

        for filename in os.listdir(self.storage_folder):
            if filename.endswith('.json'):
                filepath = os.path.join(self.storage_folder, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        submissions.append({
                            'filename': filename,
                            'filepath': filepath,
                            'enterprise_name': data.get('enterprise_info', {}).get('企业名称', ''),
                            'username': data.get('username'),  # 读取用户名
                            'submit_time': os.path.getmtime(filepath)
                        })
                except Exception as e:
                    print(f"[WARN] 读取提交记录失败: {filename} - {e}")

        # 按时间倒序排列
        submissions.sort(key=lambda x: x['submit_time'], reverse=True)

        return submissions

    def get_submission_by_filename(self, filename):
        """
        根据文件名获取提交记录

        Args:
            filename: JSON文件名

        Returns:
            提交数据
        """
        filepath = os.path.join(self.storage_folder, filename)

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"提交记录不存在: {filename}")

        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)

    def calculate_score_for_submission(self, submission_data):
        """
        计算提交问卷的得分

        Args:
            submission_data: 提交数据

        Returns:
            得分摘要
        """
        # 构造问卷数据格式
        questionnaire_data = {
            'enterprise_info': submission_data['enterprise_info'],
            'answers': []
        }

        # 读取指标体系
        indicators_df = pd.read_excel('指标体系.xlsx')
        answers = submission_data['answers']

        for idx, row in indicators_df.iterrows():
            sequence = row.get('序号', idx + 1)
            answer = answers.get(str(sequence), '未填写')

            questionnaire_data['answers'].append({
                'sequence': sequence,
                'level1': row.get('一级指标', ''),
                'level2': row.get('二级指标', ''),
                'question': row.get('三级指标（问题）', ''),
                'question_type': row.get('问题类型', ''),
                'base_score': row.get('分值', 0),
                'answer': answer,
                'applicable_enterprises': row.get('适用企业类型', '所有企业')
            })

        # 计算得分
        score_summary = self.calculator.calculate_total_score(questionnaire_data)

        return score_summary


if __name__ == '__main__':
    # 测试
    manager = QuestionnaireSubmissionManager()
    print("[INFO] 问卷提交管理器已初始化")

    # 查看所有提交
    submissions = manager.get_all_submissions()
    print(f"\n[INFO] 共有 {len(submissions)} 条提交记录")

    for submission in submissions[:5]:  # 只显示前5条
        print(f"  - {submission['enterprise_name']} ({submission['filename']})")
